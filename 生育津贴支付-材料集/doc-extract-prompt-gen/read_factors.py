#!/usr/bin/env python3
import json
import sys
from pathlib import Path

import openpyxl

HEADER_ITEM_NAME = "\u4e8b\u9879\u540d\u79f0"
HEADER_MATERIAL_NAME = "\u6750\u6599\u540d\u79f0"
HEADER_FACTOR_FIELD_NAME = "\u8981\u7d20\u5b57\u6bb5\u540d\u79f0"
HEADER_FACTOR_NAME = "\u8981\u7d20\u540d\u79f0"
HEADER_FACTOR_DESC = "\u8981\u7d20\u63d0\u53d6\u8bf4\u660e"


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def has_any_value(row):
    return any(text(value) for value in row)


def find_column(headers, aliases):
    for alias in aliases:
        for index, header in enumerate(headers):
            if header == alias:
                return index
    return None


def looks_like_material_name(value):
    return bool(value) and "\n" not in value and len(value) <= 80


def parse_factors(excel_path):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []

        headers = [text(value) for value in header_row]
        material_idx = find_column(headers, [HEADER_MATERIAL_NAME])
        extended_factor_idx = find_column(headers, [HEADER_FACTOR_FIELD_NAME])
        simple_factor_idx = find_column(headers, [HEADER_FACTOR_NAME])
        description_idx = find_column(headers, [HEADER_FACTOR_DESC])
        item_idx = find_column(headers, [HEADER_ITEM_NAME])

        is_extended = (
            extended_factor_idx is not None
            and material_idx is not None
            and (item_idx is not None or extended_factor_idx > material_idx)
        )

        factor_idx = extended_factor_idx if is_extended else simple_factor_idx
        if factor_idx is None:
            factor_idx = extended_factor_idx
        if factor_idx is None:
            raise ValueError("missing factor column in factors.xlsx")

        current_material = ""
        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not has_any_value(row):
                continue

            if material_idx is not None and material_idx < len(row):
                candidate = text(row[material_idx])
                if looks_like_material_name(candidate):
                    current_material = candidate

            factor_name = text(row[factor_idx]) if factor_idx < len(row) else ""
            if not factor_name or "\n" in factor_name or len(factor_name) > 80:
                continue

            description = ""
            if description_idx is not None and description_idx < len(row):
                description = text(row[description_idx])

            results.append(
                {
                    "field_name": factor_name,
                    "field_code": "",
                    "description": description,
                    "required": True,
                    "data_type": "string",
                    "material": current_material,
                }
            )

        return results
    finally:
        workbook.close()


def main():
    if len(sys.argv) < 2:
        print("usage: read_factors.py <factors.xlsx>", file=sys.stderr)
        return 2

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"file not found: {excel_path}", file=sys.stderr)
        return 2

    try:
        results = parse_factors(excel_path)
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1

    sys.stdout.write(json.dumps(results, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
