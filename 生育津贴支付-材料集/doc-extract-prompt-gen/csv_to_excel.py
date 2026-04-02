#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV 转 Excel 工具
将 factors.csv 转换为 factors.xlsx
"""

import os
import sys
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("安装命令: pip install openpyxl")
    sys.exit(1)

def csv_to_excel(csv_path, excel_path=None):
    """将 CSV 文件转换为 Excel 文件"""
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"未找到CSV文件: {csv_path}")
    
    if excel_path is None:
        excel_path = csv_path.replace('.csv', '.xlsx')
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "要素列表"
    
    # 读取 CSV
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # 写入 Excel
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 设置表头样式
            if row_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大宽度50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存
    wb.save(excel_path)
    print(f"✓ 成功转换: {csv_path} -> {excel_path}")
    return excel_path

def main():
    if len(sys.argv) > 1:
        csv_path = sys.argv[1]
    else:
        csv_path = 'factors.csv'
    
    try:
        excel_path = csv_to_excel(csv_path)
        print(f"\n转换完成！")
        print(f"Excel 文件: {excel_path}")
    except Exception as e:
        print(f"错误: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
