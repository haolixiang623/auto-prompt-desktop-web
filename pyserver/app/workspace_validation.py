from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from .review_rule_placeholders import (
    extract_review_rule_placeholders,
    resolve_review_rule_placeholder_token,
)
from .review_rule_builtin_variables import (
    load_review_rule_builtin_variables,
    map_review_rule_builtin_variables_by_token,
)


MEDIA_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp", ".pdf"}
LEGACY_NON_MATERIAL_DIRS = {"待分类材料", "待分类", "已分类材料"}


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_index(headers: list[str], aliases: tuple[str, ...]) -> int:
    for alias in aliases:
        for index, header in enumerate(headers):
            if header == alias or alias in header:
                return index
    return -1


def _looks_like_material_name(value: str) -> bool:
    return bool(value) and "\n" not in value and len(value) <= 80


def _push_unique(messages: list[str], message: str) -> None:
    if message not in messages:
        messages.append(message)


def _push_error(
    errors: list[str],
    diagnostics: list[dict[str, Any]],
    message: str,
    **detail: Any,
) -> None:
    if message in errors:
        return
    errors.append(message)
    diagnostic = {"level": "error", "message": message}
    for key, value in detail.items():
        if value is None or value == "":
            continue
        diagnostic[key] = value
    diagnostics.append(diagnostic)


def collect_workspace_material_dirs(root: Path) -> list[dict[str, Any]]:
    materials: list[dict[str, Any]] = []
    if not root.exists() or not root.is_dir():
        return materials

    for entry in sorted(root.iterdir(), key=lambda item: item.name):
        if not entry.is_dir():
            continue
        if entry.name.startswith(".") or entry.name.startswith("__"):
            continue
        if entry.name in LEGACY_NON_MATERIAL_DIRS:
            continue
        files = [
            child
            for child in sorted(entry.iterdir(), key=lambda item: item.name)
            if child.is_file() and child.suffix.lower() in MEDIA_EXTENSIONS
        ]
        if not files:
            continue
        materials.append(
            {
                "name": entry.name,
                "path": str(entry),
                "image_count": len(files),
                "files": files,
            }
        )
    return materials


def collect_workspace_sample_files(root: Path) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for material in collect_workspace_material_dirs(root):
        for file_path in material["files"]:
            result.append(
                {
                    "name": file_path.name,
                    "path": str(file_path),
                    "size": file_path.stat().st_size,
                    "material": material["name"],
                }
            )
    return result


def validate_workspace_bundle(paths: Any, work_dir: str, selected_materials: list[str] | None = None) -> dict[str, Any]:
    selected_materials = selected_materials or []
    root = Path(work_dir)
    errors: list[str] = []
    diagnostics: list[dict[str, Any]] = []
    warnings: list[str] = []
    builtin_variables = load_review_rule_builtin_variables(paths)
    builtin_variable_map = map_review_rule_builtin_variables_by_token(builtin_variables)
    meta: dict[str, Any] = {
        "work_dir": str(root),
        "factors_file": "",
        "headers": [],
        "material_count": 0,
        "factor_count": 0,
        "keypoint_count": 0,
        "sample_file_count": 0,
        "materials": [],
        "review_rule_builtin_variables": builtin_variables,
    }

    if not root.exists() or not root.is_dir():
        message = f"工作区不存在或不可访问: {work_dir}"
        _push_error(errors, diagnostics, message, code="workspace_not_found", path=str(root))
        return {"ok": False, "errors": errors, "warnings": [], "diagnostics": diagnostics, "meta": meta, "stats": meta}

    factors_path = root / "factors.xlsx"
    if not factors_path.exists():
        legacy_candidates = [root / name for name in ("factors.xls", "factors.csv") if (root / name).exists()]
        if legacy_candidates:
            _push_error(
                errors,
                diagnostics,
                f"当前仅检测到 {legacy_candidates[0].name}，统一校验要求工作区根目录提供 factors.xlsx。",
                code="missing_factors_xlsx",
                path=str(factors_path),
            )
        else:
            _push_error(
                errors,
                diagnostics,
                "未找到 factors.xlsx 文件，请确认工作区根目录存在该文件。",
                code="missing_factors_xlsx",
                path=str(factors_path),
            )
        return {"ok": False, "errors": errors, "warnings": warnings, "diagnostics": diagnostics, "meta": meta, "stats": meta}

    meta["factors_file"] = str(factors_path)

    try:
        with open(factors_path, "rb") as handle:
            magic = handle.read(4)
        if magic[:2] != b"PK":
            _push_error(
                errors,
                diagnostics,
                f"factors.xlsx 文件格式异常，当前文件可能不是有效的 Excel：{factors_path.name}",
                code="invalid_factors_file",
                path=str(factors_path),
            )
            return {"ok": False, "errors": errors, "warnings": warnings, "diagnostics": diagnostics, "meta": meta, "stats": meta}

        workbook = openpyxl.load_workbook(factors_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [_text(value) for value in (header_row or [])]
            meta["headers"] = headers

            material_idx = _header_index(headers, ("材料名称",))
            factor_idx = _header_index(headers, ("要素字段名称", "要素名称"))
            keypoint_idx = _header_index(headers, ("审查要点名称",))
            rule_desc_idx = _header_index(headers, ("审查要点规则说明",))

            if material_idx < 0:
                _push_error(errors, diagnostics, "factors.xlsx 缺少统一校验必需列：材料名称。", code="missing_required_column", column="材料名称")
            if factor_idx < 0:
                _push_error(errors, diagnostics, "factors.xlsx 缺少统一校验必需列：要素字段名称（或要素名称）。", code="missing_required_column", column="要素字段名称")
            if keypoint_idx < 0:
                _push_error(errors, diagnostics, "factors.xlsx 缺少统一校验必需列：审查要点名称。", code="missing_required_column", column="审查要点名称")
            if rule_desc_idx < 0:
                _push_error(errors, diagnostics, "factors.xlsx 缺少统一校验必需列：审查要点规则说明。", code="missing_required_column", column="审查要点规则说明")

            material_dirs = collect_workspace_material_dirs(root)
            sample_files = collect_workspace_sample_files(root)
            meta["materials"] = [
                {
                    "name": item["name"],
                    "path": item["path"],
                    "image_count": item["image_count"],
                }
                for item in material_dirs
            ]
            meta["material_count"] = len(material_dirs)
            meta["sample_file_count"] = len(sample_files)

            if not material_dirs:
                _push_error(
                    errors,
                    diagnostics,
                    "工作区未检测到任何材料目录，统一结构要求使用“根目录 + 一级材料目录 + 目录内图片/PDF附件”。",
                    code="missing_material_directories",
                )

            if errors:
                return {"ok": False, "errors": errors, "warnings": warnings, "diagnostics": diagnostics, "meta": meta, "stats": meta}

            current_material = ""
            factor_counts: dict[tuple[str, str], int] = defaultdict(int)
            factor_rows: dict[tuple[str, str], list[int]] = defaultdict(list)
            factor_names_by_material: dict[str, set[str]] = defaultdict(set)
            keypoint_refs_to_validate: list[dict[str, Any]] = []
            keypoint_count = 0

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue

                if material_idx < len(row):
                    candidate = _text(row[material_idx])
                    if _looks_like_material_name(candidate):
                        current_material = candidate

                if factor_idx < len(row):
                    factor_name = _text(row[factor_idx])
                    if factor_name:
                        if not current_material:
                            _push_error(
                                errors,
                                diagnostics,
                                f"第 {row_number} 行要素「{factor_name}」前缺少材料名称。",
                                code="missing_material_before_factor",
                                row=row_number,
                                column="材料名称",
                                factorName=factor_name,
                            )
                        else:
                            factor_counts[(current_material, factor_name)] += 1
                            factor_rows[(current_material, factor_name)].append(row_number)
                            factor_names_by_material[current_material].add(factor_name)

                kpname = _text(row[keypoint_idx]) if keypoint_idx < len(row) else ""
                if not kpname:
                    continue

                keypoint_count += 1
                if not current_material:
                    _push_error(
                        errors,
                        diagnostics,
                        f"第 {row_number} 行审查要点「{kpname}」前缺少材料名称。",
                        code="missing_material_before_keypoint",
                        row=row_number,
                        column="材料名称",
                        keypointName=kpname,
                    )
                    continue

                rule_desc = _text(row[rule_desc_idx]) if rule_desc_idx < len(row) else ""
                if not rule_desc:
                    _push_error(
                        errors,
                        diagnostics,
                        f"第 {row_number} 行审查要点「{kpname}」的审查要点规则说明为空。",
                        code="empty_rule_description",
                        row=row_number,
                        column="审查要点规则说明",
                        keypointName=kpname,
                    )
                    continue

                placeholders = extract_review_rule_placeholders(rule_desc)

                if not placeholders:
                    keypoint_refs_to_validate.append(
                        {
                            "row_number": row_number,
                            "kpname": kpname,
                            "material_name": current_material,
                            "placeholders": [],
                        }
                    )
                    continue

                keypoint_refs_to_validate.append(
                    {
                        "row_number": row_number,
                        "kpname": kpname,
                        "material_name": current_material,
                        "placeholders": placeholders,
                    }
                )

            meta["factor_count"] = sum(len(names) for names in factor_names_by_material.values())
            meta["keypoint_count"] = keypoint_count

            if meta["factor_count"] == 0:
                _push_error(errors, diagnostics, "factors.xlsx 未解析出任何要素定义。", code="missing_factor_definitions")
            if keypoint_count == 0:
                _push_error(
                    errors,
                    diagnostics,
                    "未读取到任何审查要点信息（审查要点名称列为空）。",
                    code="missing_keypoints",
                    column="审查要点名称",
                )

            for (material_name, factor_name), count in sorted(factor_counts.items()):
                if count > 1:
                    _push_error(
                        errors,
                        diagnostics,
                        f"材料「{material_name}」下要素「{factor_name}」重复出现 {count} 次，请在 factors.xlsx 中去重。",
                        code="duplicate_factor",
                        column="要素字段名称",
                        materialName=material_name,
                        factorName=factor_name,
                        rowNumbers=factor_rows.get((material_name, factor_name), []),
                    )

            factor_material_names = set(factor_names_by_material.keys())
            dir_material_names = {item["name"] for item in material_dirs}

            for material_name in sorted(factor_material_names - dir_material_names):
                _push_error(
                    errors,
                    diagnostics,
                    f"材料「{material_name}」在 factors.xlsx 中存在要素定义，但工作区缺少同名材料目录或目录内无可用附件。",
                    code="missing_material_directory",
                    materialName=material_name,
                )

            for material_name in sorted(dir_material_names - factor_material_names):
                _push_error(
                    errors,
                    diagnostics,
                    f"工作区材料目录「{material_name}」存在附件，但 factors.xlsx 中没有对应要素定义。",
                    code="missing_factor_definition_for_material",
                    materialName=material_name,
                )

            for material_name in selected_materials:
                if material_name not in dir_material_names:
                    _push_error(
                        errors,
                        diagnostics,
                        f"选中材料「{material_name}」在工作区中不存在同名材料目录。",
                        code="selected_material_missing_directory",
                        materialName=material_name,
                    )
                if material_name not in factor_material_names:
                    _push_error(
                        errors,
                        diagnostics,
                        f"选中材料「{material_name}」在 factors.xlsx 中没有对应要素定义。",
                        code="selected_material_missing_factors",
                        materialName=material_name,
                    )

            for item in keypoint_refs_to_validate:
                resolved_refs: list[tuple[str, str]] = []
                for token in item["placeholders"]:
                    resolved = resolve_review_rule_placeholder_token(
                        token,
                        current_material=item["material_name"],
                        builtin_variable_map=builtin_variable_map,
                        known_factor_names=factor_names_by_material.get(item["material_name"], set()),
                    )
                    if resolved["kind"] == "factor":
                        resolved_refs.append((resolved["material"], resolved["field"]))
                        continue
                    if resolved["kind"] == "builtin":
                        continue
                    _push_error(
                        errors,
                        diagnostics,
                        f"第 {item['row_number']} 行审查要点「{item['kpname']}」引用的占位符「#{token}#」无效："
                        "请使用 #材料名称-要素名称#，或先在设置中维护该内置变量。",
                        code="invalid_placeholder",
                        row=item["row_number"],
                        column="审查要点规则说明",
                        keypointName=item["kpname"],
                        token=token,
                    )

                for ref_material, ref_factor in resolved_refs:
                    if ref_material not in factor_names_by_material:
                        _push_error(
                            errors,
                            diagnostics,
                            f"第 {item['row_number']} 行审查要点「{item['kpname']}」引用的材料「{ref_material}」不存在对应要素定义。",
                            code="missing_referenced_material",
                            row=item["row_number"],
                            column="审查要点规则说明",
                            keypointName=item["kpname"],
                            materialName=ref_material,
                        )
                        continue
                    if ref_factor not in factor_names_by_material[ref_material]:
                        _push_error(
                            errors,
                            diagnostics,
                            f"第 {item['row_number']} 行审查要点「{item['kpname']}」引用的要素「{ref_material}-{ref_factor}」不存在，请先在 factors.xlsx 中补充该要素。",
                            code="missing_referenced_factor",
                            row=item["row_number"],
                            column="审查要点规则说明",
                            keypointName=item["kpname"],
                            materialName=ref_material,
                            factorName=ref_factor,
                            token=f"{ref_material}-{ref_factor}",
                        )

            return {
                "ok": len(errors) == 0,
                "errors": errors,
                "warnings": warnings,
                "diagnostics": diagnostics,
                "meta": meta,
                "stats": meta,
            }
        finally:
            workbook.close()
    except Exception as exc:
        _push_error(errors, diagnostics, f"factors.xlsx 解析失败：{exc}", code="parse_failed", path=str(factors_path))
        return {"ok": False, "errors": errors, "warnings": warnings, "diagnostics": diagnostics, "meta": meta, "stats": meta}
