from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_FACTORS_HEADERS = [
    "材料名称",
    "要素字段名称",
    "要素提取说明",
    "审查要点名称",
    "审查要点规则说明",
]
DISPLAY_CARRY_FORWARD_HEADER_PATTERNS = ["事项名称", "材料名称"]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_values(values: list[Any], size: int) -> list[str]:
    normalized = [_text(value) for value in list(values or [])[:size]]
    if len(normalized) < size:
        normalized.extend([""] * (size - len(normalized)))
    return normalized


def _empty_workbook_payload(file_path: Path) -> dict[str, Any]:
    return {
        "filePath": str(file_path),
        "exists": file_path.exists(),
        "sheetName": "Sheet1",
        "headers": list(DEFAULT_FACTORS_HEADERS),
        "rows": [],
        "mergedRanges": [],
        "summary": {
            "columnCount": len(DEFAULT_FACTORS_HEADERS),
            "rowCount": 0,
            "mergedRangeCount": 0,
        },
    }


def _normalize_merged_ranges(merged_ranges: list[dict[str, Any]] | None) -> list[dict[str, int]]:
    normalized: list[dict[str, int]] = []
    seen: set[tuple[int, int, int, int]] = set()
    for item in merged_ranges or []:
        try:
            start_row = int(item.get("startRow"))
            end_row = int(item.get("endRow"))
            start_column = int(item.get("startColumn"))
            end_column = int(item.get("endColumn"))
        except (TypeError, ValueError, AttributeError):
            continue
        if start_row < 1 or end_row < start_row or start_column < 1 or end_column < start_column:
            continue
        if start_row == end_row and start_column == end_column:
            continue
        key = (start_row, end_row, start_column, end_column)
        if key in seen:
            continue
        seen.add(key)
        normalized.append(
            {
                "startRow": start_row,
                "endRow": end_row,
                "startColumn": start_column,
                "endColumn": end_column,
            }
        )
    normalized.sort(key=lambda item: (item["startRow"], item["startColumn"], item["endRow"], item["endColumn"]))
    return normalized


def _collect_merged_ranges(worksheet: Any) -> list[dict[str, int]]:
    return _normalize_merged_ranges(
        [
            {
                "startRow": cell_range.min_row,
                "endRow": cell_range.max_row,
                "startColumn": cell_range.min_col,
                "endColumn": cell_range.max_col,
            }
            for cell_range in worksheet.merged_cells.ranges
        ]
    )


def _worksheet_grid(worksheet: Any, merged_ranges: list[dict[str, int]]) -> tuple[int, list[list[str]]]:
    max_row = max([worksheet.max_row or 1, *[item["endRow"] for item in merged_ranges]], default=1)
    max_col = max([worksheet.max_column or 1, *[item["endColumn"] for item in merged_ranges]], default=1)

    rows = [
        [_text(worksheet.cell(row=row_number, column=col_number).value) for col_number in range(1, max_col + 1)]
        for row_number in range(1, max_row + 1)
    ]

    for item in merged_ranges:
        anchor_value = rows[item["startRow"] - 1][item["startColumn"] - 1]
        for row_number in range(item["startRow"], item["endRow"] + 1):
            for col_number in range(item["startColumn"], item["endColumn"] + 1):
                rows[row_number - 1][col_number - 1] = anchor_value

    max_non_empty_col = 0
    last_relevant_row = 1
    for row_number, values in enumerate(rows, start=1):
        last_non_empty_col = 0
        for col_number, value in enumerate(values, start=1):
            if value:
                last_non_empty_col = col_number
        if last_non_empty_col > 0:
            max_non_empty_col = max(max_non_empty_col, last_non_empty_col)
            last_relevant_row = row_number

    trimmed_width = max(max_non_empty_col, 1)
    trimmed_rows = [values[:trimmed_width] for values in rows[:last_relevant_row]]
    return trimmed_width, trimmed_rows


def _normalize_payload_rows(rows: list[dict[str, Any] | list[Any]] | None) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen_row_numbers: set[int] = set()
    next_row_number = 2

    for raw_row in rows or []:
        if isinstance(raw_row, dict):
            values = raw_row.get("values") or []
            candidate_row_number = raw_row.get("rowNumber")
        elif isinstance(raw_row, list):
            values = raw_row
            candidate_row_number = None
        else:
            values = []
            candidate_row_number = None

        row_number: int | None = None
        try:
            parsed_row_number = int(candidate_row_number)
        except (TypeError, ValueError):
            parsed_row_number = 0
        if parsed_row_number >= 2 and parsed_row_number not in seen_row_numbers:
            row_number = parsed_row_number
        else:
            while next_row_number in seen_row_numbers:
                next_row_number += 1
            row_number = next_row_number

        seen_row_numbers.add(row_number)
        next_row_number = max(next_row_number, row_number + 1)
        normalized.append(
            {
                "rowNumber": row_number,
                "values": [_text(value) for value in list(values)],
            }
        )

    normalized.sort(key=lambda item: item["rowNumber"])
    return normalized


def _display_carry_forward_column_indices(headers: list[str]) -> list[int]:
    return [
        index
        for index, header in enumerate(headers)
        if any(pattern in _text(header) for pattern in DISPLAY_CARRY_FORWARD_HEADER_PATTERNS)
    ]


def _carry_forward_display_columns(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    carry_indices = _display_carry_forward_column_indices(headers)
    if not carry_indices:
        return rows

    last_values = [""] * len(headers)
    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        values = _normalize_values(row.get("values") or [], len(headers))
        next_values = list(values)
        for index in carry_indices:
            if next_values[index]:
                last_values[index] = next_values[index]
            elif last_values[index]:
                next_values[index] = last_values[index]
        normalized_rows.append(
            {
                **row,
                "values": next_values,
            }
        )
    return normalized_rows


def load_factors_workbook(file_path: str | Path) -> dict[str, Any]:
    target = Path(file_path)
    if not target.exists():
        return _empty_workbook_payload(target)

    workbook = openpyxl.load_workbook(target)
    try:
        worksheet = workbook.active
        merged_ranges = _collect_merged_ranges(worksheet)
        max_non_empty_col, raw_rows = _worksheet_grid(worksheet, merged_ranges)
        if max_non_empty_col <= 0:
            return {
                **_empty_workbook_payload(target),
                "exists": True,
                "sheetName": worksheet.title,
                "mergedRanges": merged_ranges,
                "summary": {
                    "columnCount": len(DEFAULT_FACTORS_HEADERS),
                    "rowCount": 0,
                    "mergedRangeCount": len(merged_ranges),
                },
            }

        headers = _normalize_values(raw_rows[0] if raw_rows else [], max_non_empty_col)
        rows: list[dict[str, Any]] = []
        for row_number, raw_row in enumerate(raw_rows[1:], start=2):
            values = _normalize_values(raw_row, max_non_empty_col)
            rows.append(
                {
                    "rowNumber": row_number,
                    "values": values,
                }
            )
        rows = _carry_forward_display_columns(headers, rows)

        return {
            "filePath": str(target),
            "exists": True,
            "sheetName": worksheet.title,
            "headers": headers,
            "rows": rows,
            "mergedRanges": merged_ranges,
            "summary": {
                "columnCount": len(headers),
                "rowCount": len(rows),
                "mergedRangeCount": len(merged_ranges),
            },
        }
    finally:
        workbook.close()


def save_factors_workbook(
    file_path: str | Path,
    headers: list[Any],
    rows: list[dict[str, Any] | list[Any]],
    merged_ranges: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    target = Path(file_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    normalized_rows = _normalize_payload_rows(rows)
    max_row_width = max((len(item["values"]) for item in normalized_rows), default=0)

    workbook = openpyxl.load_workbook(target) if target.exists() else openpyxl.Workbook()
    try:
        worksheet = workbook.active
        existing_merged_ranges = _collect_merged_ranges(worksheet)
        normalized_merged_ranges = _normalize_merged_ranges(merged_ranges or existing_merged_ranges)
        for merged_range in [str(item) for item in worksheet.merged_cells.ranges]:
            worksheet.unmerge_cells(merged_range)

        max_merged_col = max((item["endColumn"] for item in normalized_merged_ranges), default=0)
        max_merged_row = max((item["endRow"] for item in normalized_merged_ranges), default=0)
        column_count = max(1, len(headers or []), max_row_width, max_merged_col)
        normalized_headers = _normalize_values(list(headers or []), column_count)
        normalized_row_map = {
            item["rowNumber"]: _normalize_values(item["values"], column_count)
            for item in normalized_rows
        }
        existing_max_row = max(worksheet.max_row or 1, max(normalized_row_map.keys(), default=1), max_merged_row)
        existing_max_col = max(worksheet.max_column or 1, column_count)

        for row_number in range(1, existing_max_row + 1):
            for col_number in range(1, existing_max_col + 1):
                next_value: str | None = None
                if row_number == 1 and col_number <= column_count:
                    next_value = normalized_headers[col_number - 1] or None
                elif row_number in normalized_row_map and col_number <= column_count:
                    next_value = normalized_row_map[row_number][col_number - 1] or None
                worksheet.cell(row=row_number, column=col_number).value = next_value

        for item in normalized_merged_ranges:
            start_row = item["startRow"]
            end_row = item["endRow"]
            start_column = item["startColumn"]
            end_column = item["endColumn"]
            anchor_value = ""
            for row_number in range(start_row, end_row + 1):
                for col_number in range(start_column, end_column + 1):
                    candidate = _text(worksheet.cell(row=row_number, column=col_number).value)
                    if candidate:
                        anchor_value = candidate
                        break
                if anchor_value:
                    break

            worksheet.cell(row=start_row, column=start_column).value = anchor_value or None
            for row_number in range(start_row, end_row + 1):
                for col_number in range(start_column, end_column + 1):
                    if row_number == start_row and col_number == start_column:
                        continue
                    worksheet.cell(row=row_number, column=col_number).value = None
            worksheet.merge_cells(
                start_row=start_row,
                end_row=end_row,
                start_column=start_column,
                end_column=end_column,
            )

        workbook.save(target)
    finally:
        workbook.close()

    return load_factors_workbook(target)
