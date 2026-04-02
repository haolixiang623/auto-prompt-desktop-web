"""
解析生产环境导出的事项AI配置 Excel 文件。

Excel 包含 10 个 sheet:
  ai_task, ai_task_material, ai_carrier, ai_factor,
  ai_keypoint, ai_rule, ai_material_prompt,
  ai_prompt_group, ai_prompt_group_factor, ai_keypoint_file

提供两个主要函数:
  - parse_excel_for_cases:  提取要素提示词 → 导入提示词库 (cases 表)
  - parse_excel_for_review_rules: 提取审查规则 → 导入审查规则库 (review_rules 表)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


# ─────────────────────────── helpers ───────────────────────────

def _read_sheet(wb: openpyxl.Workbook, name: str) -> tuple[list[str], list[dict]]:
    """读取指定 sheet，返回 (headers, rows_as_dicts)。sheet 不存在时返回空。"""
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        return [], []
    headers = [str(c.value).strip() if c.value else "" for c in header_row]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        rows.append(d)
    return headers, rows


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ─────────────────────── cases (提示词库) ──────────────────────

def parse_excel_for_cases(file_path: str | Path) -> dict:
    """解析 Excel，提取要素提示词数据，用于导入提示词库 (cases 表)。

    Returns:
        {
            "item_name": str,
            "cases": [
                {
                    "item_name": str,
                    "material_name": str,
                    "factor_name": str,
                    "extraction_rule": str,   # factor_prompt
                    "metadata": { factortype, factoruse, ... }
                },
                ...
            ],
            "summary": { "material_count": int, "factor_count": int }
        }
    """
    wb = openpyxl.load_workbook(str(file_path))

    # 1. 事项名称
    _, tasks = _read_sheet(wb, "ai_task")
    item_name = _str(tasks[0].get("taskname")) if tasks else ""

    # 2. carrier → material 映射 (carrierguid → carriername)
    _, carriers = _read_sheet(wb, "ai_carrier")
    carrier_map: dict[str, str] = {}  # rowguid → carriername
    for c in carriers:
        guid = _str(c.get("rowguid"))
        cname = _str(c.get("carriername"))
        if guid and cname:
            carrier_map[guid] = cname

    # 3. 要素
    _, factors = _read_sheet(wb, "ai_factor")
    cases: list[dict] = []
    materials_seen: set[str] = set()

    for f in factors:
        factor_name = _str(f.get("factorname"))
        factor_prompt = _str(f.get("factor_prompt"))
        if not factor_name:
            continue

        carrier_guid = _str(f.get("carrierguid"))
        carrier_name = _str(f.get("carriername")) or carrier_map.get(carrier_guid, "")
        materials_seen.add(carrier_name)

        cases.append({
            "item_name": item_name,
            "material_name": carrier_name,
            "factor_name": factor_name,
            "extraction_rule": factor_prompt,
            "metadata": {
                "factortype": _str(f.get("factortype")),
                "factoruse": _str(f.get("factoruse")),
                "factor_trans": _str(f.get("factor_trans")),
                "is_usermsg": _str(f.get("is_usermsg")),
                "ordernum": _str(f.get("ordernum")),
                "remark": _str(f.get("remark")),
            },
        })

    wb.close()
    return {
        "item_name": item_name,
        "cases": cases,
        "summary": {
            "material_count": len(materials_seen),
            "factor_count": len(cases),
        },
    }


# ──────────────────── review rules (审查规则库) ────────────────

def parse_excel_for_review_rules(file_path: str | Path) -> dict:
    """解析 Excel，提取审查要点 + 审查规则数据，用于导入审查规则库。

    Returns:
        {
            "item_name": str,
            "rules": [
                {
                    "item_name": str,
                    "materialname": str,
                    "keypoints": [
                        {
                            "kpname": str,
                            "review_rule": str,       # "1"/"2"/"3"
                            "review_rule_text": str,
                            "content": str,
                            "review_conditions": dict|None,
                            "review_rule_js": str,
                            "passreason": str,
                            "nopassreason": str,
                            "is_point": str,
                            "pre_rule_enabled": int,
                            "pre_conditions": dict|None,
                        },
                        ...
                    ]
                },
                ...
            ],
            "summary": { "material_count": int, "keypoint_count": int }
        }
    """
    wb = openpyxl.load_workbook(str(file_path))

    # 1. 事项名称
    _, tasks = _read_sheet(wb, "ai_task")
    item_name = _str(tasks[0].get("taskname")) if tasks else ""

    # 2. 审查要点 (ai_keypoint)
    _, keypoints = _read_sheet(wb, "ai_keypoint")
    kp_map: dict[str, dict] = {}  # rowguid → keypoint dict
    for kp in keypoints:
        guid = _str(kp.get("rowguid"))
        if guid:
            kp_map[guid] = kp

    # 3. 审查规则 (ai_rule) — 通过 kpguid 关联到 keypoint
    _, rules = _read_sheet(wb, "ai_rule")
    rule_by_kp: dict[str, dict] = {}  # kpguid → rule dict
    for r in rules:
        kpguid = _str(r.get("kpguid"))
        if kpguid:
            rule_by_kp[kpguid] = r

    # 4. 按 materialname 分组
    from collections import OrderedDict
    grouped: OrderedDict[str, list[dict]] = OrderedDict()
    kp_count = 0

    for kp_guid, kp in kp_map.items():
        mat = _str(kp.get("materialname"))
        kpname = _str(kp.get("kpname"))
        if not kpname:
            continue

        rule = rule_by_kp.get(kp_guid, {})

        # 解析 review_conditions JSON
        raw_cond = rule.get("review_conditions")
        review_conditions = None
        if raw_cond:
            try:
                review_conditions = json.loads(str(raw_cond)) if isinstance(raw_cond, str) else raw_cond
            except (json.JSONDecodeError, TypeError):
                pass

        # 解析 pre_conditions JSON
        raw_pre = rule.get("pre_conditions")
        pre_conditions = None
        if raw_pre:
            try:
                pre_conditions = json.loads(str(raw_pre)) if isinstance(raw_pre, str) else raw_pre
            except (json.JSONDecodeError, TypeError):
                pass

        kp_entry = {
            "kpname": kpname,
            "review_rule": _str(rule.get("review_rule")) or _str(kp.get("aitype")),
            "review_rule_text": kpname,
            "content": _str(kp.get("content")),
            "review_conditions": review_conditions,
            "review_rule_js": _str(rule.get("review_rule_js")),
            "passreason": _str(kp.get("passreason")),
            "nopassreason": _str(kp.get("nopassreason")),
            "is_point": _str(rule.get("is_point")),
            "pre_rule_enabled": int(rule.get("pre_rule_enabled") or 0),
            "pre_conditions": pre_conditions,
        }

        if mat not in grouped:
            grouped[mat] = []
        grouped[mat].append(kp_entry)
        kp_count += 1

    result_rules: list[dict] = []
    for mat, kps in grouped.items():
        result_rules.append({
            "item_name": item_name,
            "materialname": mat,
            "keypoints": kps,
        })

    wb.close()
    return {
        "item_name": item_name,
        "rules": result_rules,
        "summary": {
            "material_count": len(grouped),
            "keypoint_count": kp_count,
        },
    }


# ─────────────────────── preview (统一预览) ────────────────────

def preview_excel(file_path: str | Path) -> dict:
    """快速预览 Excel 内容概要，不做导入。"""
    wb = openpyxl.load_workbook(str(file_path))
    _, tasks = _read_sheet(wb, "ai_task")
    item_name = _str(tasks[0].get("taskname")) if tasks else ""
    _, carriers = _read_sheet(wb, "ai_carrier")
    _, factors = _read_sheet(wb, "ai_factor")
    _, keypoints = _read_sheet(wb, "ai_keypoint")
    _, rules = _read_sheet(wb, "ai_rule")
    wb.close()
    return {
        "item_name": item_name,
        "carrier_count": len(carriers),
        "factor_count": len(factors),
        "keypoint_count": len(keypoints),
        "rule_count": len(rules),
        "sheets": list(wb.sheetnames) if hasattr(wb, "sheetnames") else [],
    }
