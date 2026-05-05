#!/usr/bin/env python3
"""
审查规则JSON生成器
从 factors.xlsx 的审查要点列生成符合审查规则导入规范的 JSON。

规则格式说明:
  - 审查要点规则说明列: 使用 #材料名称-字段名称# 引用要素
  - 空审查要点名称: 跳过不生成
  - 审查模式 (review_rule) 由 LLM 推理决定:
      "1" = 大模型(LLM)      -> 无法明确规则对比时
      "2" = 规则对比          -> 有明确的 #材料-字段# 引用且可识别比较关系
      "3" = Groovy脚本        -> 需要复杂算法/正则/日期计算时

用法:
  python3 generate_review_rule.py <材料集目录> [--api-key KEY] [--base-url URL] [--model MODEL]
"""

import os
import sys
import json
import re
import argparse
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pyserver.app.review_rule_placeholders import (
    extract_review_rule_refs as shared_extract_review_rule_refs,
    replace_review_rule_refs_with_placeholders as shared_replace_review_rule_refs_with_placeholders,
)

# 重定义 print 函数，强制实时刷新输出
_original_print = print
def print(*args, **kwargs):
    kwargs.setdefault('flush', True)
    _original_print(*args, **kwargs)

# DashScope 通过 OpenAI 兼容接口传递时，这些参数必须放在 extra_body 而非顶层 kwargs
_DASHSCOPE_BODY_PARAMS = {
    'enable_thinking', 'thinking_budget', 'translation_options',
    'vl_high_resolution_images', 'search_options',
}

def get_extra_params():
    """读取桌面端传入的额外模型参数，拆分为标准参数和 DashScope 专有参数(extra_body)"""
    raw = os.environ.get("GENERATE_EXTRA_PARAMS", "{}")
    try:
        all_params = json.loads(raw)
    except Exception:
        all_params = {}
    standard = {k: v for k, v in all_params.items() if k not in _DASHSCOPE_BODY_PARAMS}
    body = {k: v for k, v in all_params.items() if k in _DASHSCOPE_BODY_PARAMS}
    if body:
        standard['extra_body'] = body
    return standard


DEFAULT_REVIEW_RULE_BUILTIN_VARIABLES = [
    {
        "id": "current_date",
        "name": "当前日期",
        "token": "当前日期",
        "placeholder": "$系统变量:当前日期$",
        "dataType": "date",
        "description": "当前系统日期",
    }
]


def normalize_review_rule_builtin_variable(raw):
    if isinstance(raw, str):
        raw = {"token": raw}
    item = raw if isinstance(raw, dict) else {}
    token = str(item.get("token") or item.get("name") or "").strip() or "当前日期"
    name = str(item.get("name") or token).strip() or token
    return {
        "id": str(item.get("id") or re.sub(r"[^\w]+", "_", token.lower()).strip("_") or "builtin_variable").strip(),
        "name": name,
        "token": token,
        "placeholder": str(item.get("placeholder") or f"$系统变量:{name}$").strip() or f"$系统变量:{name}$",
        "dataType": str(item.get("dataType") or item.get("type") or "string").strip() or "string",
        "description": str(item.get("description") or "").strip(),
    }


def load_review_rule_builtin_variables():
    raw = os.environ.get("REVIEW_RULE_BUILTIN_VARIABLES", "")
    if raw:
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = None
    else:
        parsed = None
    if not isinstance(parsed, list) or not parsed:
        parsed = DEFAULT_REVIEW_RULE_BUILTIN_VARIABLES
    normalized = []
    seen = set()
    for item in parsed:
        built_in = normalize_review_rule_builtin_variable(item)
        token = built_in["token"]
        if token in seen:
            continue
        seen.add(token)
        normalized.append(built_in)
    return normalized or [normalize_review_rule_builtin_variable(DEFAULT_REVIEW_RULE_BUILTIN_VARIABLES[0])]


REVIEW_RULE_BUILTIN_VARIABLES = load_review_rule_builtin_variables()
REVIEW_RULE_BUILTIN_VARIABLE_MAP = {item["token"]: item for item in REVIEW_RULE_BUILTIN_VARIABLES}

# ─────────────────────────── Excel 读取 ─────────────────────────────

def read_review_rules_from_excel(excel_path):
    """从 factors.xlsx 读取审查要点信息。

    支持两种格式:
    - 梳理表格式 (扩展格式): 含「事项名称」「材料名称」「审查要点名称」「审查要点规则说明」等列
    - 简单格式: 含「材料名称」「审查要点名称」「审查要点规则说明」列

    Returns:
        dict: {材料名称: [(kpname, rule_desc, passreason, nopassreason, ordernum, exclude_situations), ...]}
    """
    try:
        with open(excel_path, 'rb') as f:
            magic = f.read(4)
        if magic[:2] != b'PK':
            print(f"[错误] 不支持的文件格式: {excel_path}")
            return {}

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        result = {}

        header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"[信息] 表头: {header[:14]}")

        # 定位关键列
        def find_col(keywords):
            for kw in keywords:
                for i, h in enumerate(header):
                    if kw in h:
                        return i
            return -1

        IDX_MATERIAL    = find_col(["材料名称"])
        IDX_FACTOR      = find_col(["要素字段名称", "要素名称"])
        IDX_KPNAME      = find_col(["审查要点名称"])
        IDX_RULE_DESC   = find_col(["审查要点规则说明"])
        IDX_PASS        = find_col(["审查通过标准回答"])
        IDX_NOPASS      = find_col(["审查不通过标准回答"])
        IDX_ORDERNUM    = find_col(["排序号"])
        IDX_EXCLUDE     = find_col(["排他情形", "排除情形"])
        IDX_REVIEW_RULE_SPECIAL = find_col(["审查规则特殊说明"])

        if IDX_MATERIAL < 0 or IDX_KPNAME < 0:
            print(f"[错误] 未找到必要列（材料名称/审查要点名称）")
            return {}

        print(f"[信息] 列索引 - 材料:{IDX_MATERIAL} 要点名称:{IDX_KPNAME} 规则说明:{IDX_RULE_DESC} "
              f"通过:{IDX_PASS} 不通过:{IDX_NOPASS}")

        current_material = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 继承材料名称（合并单元格向下继承）
            raw_mat = row[IDX_MATERIAL] if IDX_MATERIAL < len(row) else None
            if raw_mat:
                s = str(raw_mat).strip()
                if s and '\n' not in s and len(s) < 80:
                    current_material = s

            if not current_material:
                continue

            kpname = ""
            if IDX_KPNAME >= 0 and IDX_KPNAME < len(row) and row[IDX_KPNAME]:
                kpname = str(row[IDX_KPNAME]).strip()

            if not kpname:
                continue

            factor_name = ""
            if IDX_FACTOR >= 0 and IDX_FACTOR < len(row) and row[IDX_FACTOR]:
                factor_name = str(row[IDX_FACTOR]).strip()

            rule_desc = ""
            if IDX_RULE_DESC >= 0 and IDX_RULE_DESC < len(row) and row[IDX_RULE_DESC]:
                rule_desc = str(row[IDX_RULE_DESC]).strip()

            passreason = ""
            if IDX_PASS >= 0 and IDX_PASS < len(row) and row[IDX_PASS]:
                passreason = str(row[IDX_PASS]).strip()

            nopassreason = ""
            if IDX_NOPASS >= 0 and IDX_NOPASS < len(row) and row[IDX_NOPASS]:
                nopassreason = str(row[IDX_NOPASS]).strip()

            ordernum = None
            if IDX_ORDERNUM >= 0 and IDX_ORDERNUM < len(row) and row[IDX_ORDERNUM]:
                try:
                    ordernum = int(row[IDX_ORDERNUM])
                except Exception:
                    pass

            exclude_situations = ""
            if IDX_EXCLUDE >= 0 and IDX_EXCLUDE < len(row) and row[IDX_EXCLUDE]:
                exclude_situations = str(row[IDX_EXCLUDE]).strip()

            special_note = ""
            if IDX_REVIEW_RULE_SPECIAL >= 0 and IDX_REVIEW_RULE_SPECIAL < len(row) and row[IDX_REVIEW_RULE_SPECIAL]:
                special_note = str(row[IDX_REVIEW_RULE_SPECIAL]).strip()

            result.setdefault(current_material, [])
            result[current_material].append({
                "factor_name": factor_name,
                "kpname": kpname,
                "rule_desc": rule_desc,
                "passreason": passreason,
                "nopassreason": nopassreason,
                "ordernum": ordernum,
                "exclude_situations": exclude_situations,
                "special_note": special_note,
            })

        return result

    except Exception as e:
        print(f"[错误] 读取Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return {}


# ─────────────────────────── 要素引用解析 ─────────────────────────────

def is_builtin_variable_token(token):
    return str(token or "").strip() in REVIEW_RULE_BUILTIN_VARIABLE_MAP


def get_builtin_variable(token):
    return REVIEW_RULE_BUILTIN_VARIABLE_MAP.get(str(token or "").strip())


def extract_review_refs(text, current_material="", known_factor_names=None):
    """按出现顺序提取规则中的引用，支持 factor 和 builtin variable。"""
    return shared_extract_review_rule_refs(
        text,
        current_material=current_material,
        builtin_variable_map=REVIEW_RULE_BUILTIN_VARIABLE_MAP,
        known_factor_names=known_factor_names,
    )


def extract_factor_refs(text, current_material="", known_factor_names=None):
    """从 #材料-字段# 格式中提取 [(材料名称, 字段名称), ...] 列表"""
    return [
        (ref["material"], ref["field"])
        for ref in extract_review_refs(text, current_material=current_material, known_factor_names=known_factor_names)
        if ref["kind"] == "factor"
    ]


def extract_builtin_refs(text, current_material="", known_factor_names=None):
    return [
        ref
        for ref in extract_review_refs(text, current_material=current_material, known_factor_names=known_factor_names)
        if ref["kind"] == "builtin"
    ]


def refs_to_factor_placeholders(refs):
    """将要素引用列表转为 $载体:要素$ 格式"""
    return [make_factor_placeholder(mat, field) for mat, field in refs]


def make_factor_placeholder(material_name, field_name):
    """仅规范中间分隔符为英文半角冒号，保留两侧原文"""
    mat = str(material_name or "").replace("：", ":").strip()
    field = str(field_name or "").replace("：", ":").strip()
    return f"${mat}:{field}$"


def replace_rule_refs_with_placeholders(text, current_material="", known_factor_names=None):
    return shared_replace_review_rule_refs_with_placeholders(
        text,
        current_material=current_material,
        builtin_variable_map=REVIEW_RULE_BUILTIN_VARIABLE_MAP,
        known_factor_names=known_factor_names,
    )


# ─────────────────────────── 规则类型推断 ─────────────────────────────

def infer_review_rule_local(rule_desc, special_note="", factor_name="", material_name="", known_factor_names=None):
    """本地推断审查规则类型 (不依赖LLM)。

    推断逻辑:
    1. 若含 #材料-字段# 引用 -> 尝试识别为 review_rule="2" (规则对比)
       - 含比较词（一致/相同/等于/不超过/大于/小于/包含 等）-> "2"
       - 含日期比较/有效期 -> "2"
       - 含复杂算法关键词（计算/天数/月份/工作日 等）-> "3"
    2. 无要素引用但有 #材料-字段# -> "1" (大模型)
    3. 无任何引用 -> "1"

    Returns:
        str: "1", "2", or "3"
    """
    combined = (rule_desc or "") + " " + (special_note or "")

    refs = extract_review_refs(
        combined,
        current_material=material_name,
        known_factor_names=known_factor_names,
    )
    factor_refs = [ref for ref in refs if ref["kind"] == "factor"]
    builtin_refs = [ref for ref in refs if ref["kind"] == "builtin"]

    # 含引用时进一步判断
    if factor_refs or builtin_refs:
        groovy_keywords = ["计算", "天数", "月份", "工作日", "算法", "正则", "公式", "统计", "求和", "截取", "substring"]
        for kw in groovy_keywords:
            if kw in combined:
                return "3"

        comparison_keywords = [
            "一致", "相同", "等于", "须与", "需与", "应与", "应为", "须为", "不超过",
            "大于", "小于", "不小于", "不大于", "至少", "最多", "包含", "有效期", "在有效期",
            "是否过期", "不能为空", "非空", "长度", "格式", "规范"
        ]
        for kw in comparison_keywords:
            if kw in combined:
                return "2"

        if factor_refs:
            return "2"
        if builtin_refs and factor_name:
            return "2"
        return "1"

    # 无引用 -> 大模型
    return "1"


def infer_operator_and_type(rule_desc, ref_a, ref_b=None):
    """根据规则描述推断运算符和数据类型。

    Returns:
        (operator, data_type, string_replacements, delimiter)
    """
    text = rule_desc or ""

    # 日期相关
    if any(kw in text for kw in ["有效期", "日期", "到期", "过期", "截止"]):
        return "ge", "date", None, None

    # 数值相关 (含金额/数量)
    if any(kw in text for kw in ["注册资本", "金额", "数量", "个数", "万元", "元"]):
        sr = None
        if "万元" in text or "元" in text:
            sr = [{"old_value": "万元", "new_value": ""}, {"old_value": "元", "new_value": ""}]
        if any(kw in text for kw in ["大于", "超过", "不小于", "至少"]):
            return "ge", "float", sr, None
        if any(kw in text for kw in ["小于", "不超过", "最多"]):
            return "le", "float", sr, None
        return "eq", "float", sr, None

    # 包含关系
    if any(kw in text for kw in ["包含", "涵盖", "含有"]):
        delimiter = None
        if "、" in text or "顿号" in text:
            delimiter = "、"
        return "contains", "string", None, delimiter

    # 长度校验
    if "18位" in text or "位数" in text:
        return "len_eq", "string", None, None

    # 非空检查
    if any(kw in text for kw in ["非空", "不能为空", "不为空", "必须填写", "已识别"]):
        return "notblank", "string", None, None

    # 默认: 等值比较
    return "eq", "string", None, None


# ─────────────────────────── LLM 推理 (可选) ─────────────────────────────

def call_llm_for_rule(kpname, rule_desc, material_name, api_key, base_url, model, timeout=120):
    """调用LLM分析审查规则，返回结构化JSON。

    Returns:
        dict or None
    """
    import time
    start_time = time.time()
    try:
        import urllib.request
        import urllib.error

        builtin_lines = "\n".join(
            f"- `#{item['token']}#` 表示内置变量，转换时使用 `{item['placeholder']}`，比较对象类型填 `variable`"
            for item in REVIEW_RULE_BUILTIN_VARIABLES
        )

        prompt = f"""你是一个审查规则分析专家。请根据以下审查要点规则说明，生成符合导入规范的审查规则JSON。

## 审查背景
- 材料名称: {material_name}
- 审查要点名称: {kpname}
- 审查要点规则说明: {rule_desc}

## 规则格式说明
- `#材料名称-字段名称#` 表示引用该材料下的某个要素，转换时用 `$材料名称:字段名称$` 格式
- `#字段名称#` 表示当前材料下该字段的简写引用，例如材料名称为“营业证照”时，`#统一社会信用代码#` 等价于 `#营业证照-统一社会信用代码#`
- 内置变量说明：
{builtin_lines or '- 当前未配置内置变量'}
- review_rule: "1"=大模型(LLM), "2"=规则对比, "3"=Groovy脚本

## 判断逻辑
1. 若规则说明中有明确的字段比较（如"A需与B一致"，"A不超过B"），选 review_rule="2"
2. 若需要复杂算法（计算天数、正则校验格式等），选 review_rule="3"  
3. 若规则模糊或为文字描述，选 review_rule="1"

## 输出格式（严格JSON，无多余内容）
{{
  "review_rule": "1|2|3",
  "review_rule_text": "简洁的审查规则文本描述",
  "content": "当review_rule=1时，填写LLM提示词；否则为空",
  "passreason": "审查通过原因模板",
  "nopassreason": "审查不通过原因模板",
  "review_conditions": null,
  "review_rule_js": ""
}}

当 review_rule="2" 时，review_conditions 格式:
{{
  "groups": [
    {{
      "logicToNext": null,
      "groupFailReason": "不通过原因，用$材料:字段$占位符",
      "conditions": [
        {{
          "elementA": "$材料A:字段A$",
          "elementAType": "factor",
          "elementADisplay": "$材料A:字段A$",
          "operator": "eq|nq|gt|lt|ge|le|contains|notblank|blank|len_gt|len_lt|len_eq|regex",
          "dataType": "string|int|float|date|array",
          "elementB": "$材料B:字段B$或固定值",
          "elementBType": "factor|value|variable",
          "elementBDisplay": "$材料B:字段B$或固定值",
          "logicToNext": null,
          "stringReplacements": null,
          "delimiter": null,
          "arrayKeys": null
        }}
      ]
    }}
  ]
}}

当 review_rule="3" 时，review_rule_js 为 Groovy 脚本，使用 input.get("材料:字段") 获取要素值。

请直接输出JSON，不要有其他内容。"""

        # 获取额外参数（如 enable_thinking）
        extra_params = get_extra_params()

        payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens": 2000,
        }
        # 合并额外参数
        payload.update(extra_params)
        data = json.dumps(payload).encode('utf-8')

        req = urllib.request.Request(
            f"{base_url.rstrip('/')}/chat/completions",
            data=data,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            method="POST"
        )

        with urllib.request.urlopen(req, timeout=timeout) as resp:
            resp_data = json.loads(resp.read().decode('utf-8'))

        content = resp_data["choices"][0]["message"]["content"].strip()
        elapsed = time.time() - start_time

        # 输出 LLM 调用日志
        log_entry = {
            "model": model,
            "scene": "审查规则生成",
            "prompt_summary": prompt[:2000],
            "response_summary": content[:2000],
            "elapsed_s": elapsed,
            "success": True,
        }
        print(f"__LLM_LOG__:{json.dumps(log_entry, ensure_ascii=False)}", file=sys.stderr, flush=True)

        # 提取JSON部分
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
        return None

    except Exception as e:
        elapsed = time.time() - start_time
        # 输出 LLM 调用日志（失败）
        log_entry = {
            "model": model,
            "scene": "审查规则生成",
            "prompt_summary": f"要点: {kpname}, 规则: {rule_desc[:100]}",
            "response_summary": "",
            "elapsed_s": elapsed,
            "success": False,
            "error": str(e),
        }
        print(f"__LLM_LOG__:{json.dumps(log_entry, ensure_ascii=False)}", file=sys.stderr, flush=True)
        print(f"  [警告] LLM调用失败: {e}，使用本地推断")
        return None


# ─────────────────────────── 规则构建 ─────────────────────────────

def build_keypoint_rule2(kpname, rule_desc, passreason, nopassreason, material_name, ordernum, exclude_situations, factor_name="", known_factor_names=None):
    """构建 review_rule="2" (规则对比) 的要点JSON。"""
    refs = [
        ref
        for ref in extract_review_refs(
            rule_desc,
            current_material=material_name,
            known_factor_names=known_factor_names,
        )
        if ref["kind"] != "invalid"
    ]
    factor_refs = [ref for ref in refs if ref["kind"] == "factor"]
    builtin_refs = [ref for ref in refs if ref["kind"] == "builtin"]

    if not refs:
        # 无引用但被判断为2 -> 降级为1
        return build_keypoint_rule1(
            kpname,
            rule_desc,
            passreason,
            nopassreason,
            material_name,
            ordernum,
            exclude_situations,
            factor_name=factor_name,
            known_factor_names=known_factor_names,
        )

    def current_factor_ref():
        if not factor_name:
            return None
        return {
            "kind": "factor",
            "token": f"{material_name}-{factor_name}",
            "material": material_name,
            "field": factor_name,
        }

    def placeholder_of(ref):
        if ref["kind"] == "builtin":
            return ref["placeholder"]
        return make_factor_placeholder(ref["material"], ref["field"])

    def element_type_of(ref):
        if ref["kind"] == "builtin":
            return "variable"
        system_carriers = {"常规信息", "法人信息", "自然人信息"}
        return "variable" if ref["material"] in system_carriers else "factor"

    def tuple_ref(ref):
        if ref["kind"] == "builtin":
            return ("系统变量", ref["name"])
        return (ref["material"], ref["field"])

    # 分析比较关系
    conditions = []
    groups = []

    if len(refs) == 1 and refs[0]["kind"] == "factor":
        # 单要素: 非空检查
        mat, field = refs[0]["material"], refs[0]["field"]
        placeholder = make_factor_placeholder(mat, field)
        operator, data_type, sr, delimiter = infer_operator_and_type(rule_desc, (mat, field))

        if operator in ("notblank", "blank"):
            cond = {
                "elementA": placeholder,
                "elementAType": "factor",
                "elementADisplay": placeholder,
                "operator": operator,
                "dataType": data_type,
                "elementB": "",
                "elementBType": "value",
                "elementBDisplay": "",
                "logicToNext": None,
                "stringReplacements": None,
                "delimiter": None,
                "arrayKeys": None,
            }
            group_fail = nopassreason or ""
        else:
            # 尝试找固定值比较 (如 "长度为18位")
            len_match = re.search(r'(\d+)位', rule_desc)
            fixed_val = ""
            if len_match and operator in ("len_eq", "len_gt", "len_lt"):
                fixed_val = len_match.group(1)
            cond = {
                "elementA": placeholder,
                "elementAType": "factor",
                "elementADisplay": placeholder,
                "operator": operator,
                "dataType": data_type,
                "elementB": fixed_val,
                "elementBType": "value",
                "elementBDisplay": fixed_val,
                "logicToNext": None,
                "stringReplacements": sr,
                "delimiter": delimiter,
                "arrayKeys": None,
            }
            group_fail = nopassreason or ""

        conditions.append(cond)
        groups.append({
            "logicToNext": None,
            "groupFailReason": group_fail,
            "conditions": conditions,
        })

    elif len(factor_refs) == 1 and len(builtin_refs) == 1:
        factor_ref = factor_refs[0]
        builtin_ref = builtin_refs[0]
        placeholder_a = placeholder_of(factor_ref)
        placeholder_b = placeholder_of(builtin_ref)

        operator, data_type, sr, delimiter = infer_operator_and_type(rule_desc, tuple_ref(factor_ref), tuple_ref(builtin_ref))

        cond = {
            "elementA": placeholder_a,
            "elementAType": "factor",
            "elementADisplay": placeholder_a,
            "operator": operator,
            "dataType": data_type,
            "elementB": placeholder_b,
            "elementBType": "variable",
            "elementBDisplay": placeholder_b,
            "logicToNext": None,
            "stringReplacements": sr,
            "delimiter": delimiter,
            "arrayKeys": None,
        }
        conditions.append(cond)
        groups.append({
            "logicToNext": None,
            "groupFailReason": nopassreason or "",
            "conditions": conditions,
        })

    elif len(factor_refs) == 0 and len(builtin_refs) == 1 and current_factor_ref():
        factor_ref = current_factor_ref()
        builtin_ref = builtin_refs[0]
        placeholder_a = placeholder_of(factor_ref)
        placeholder_b = placeholder_of(builtin_ref)

        operator, data_type, sr, delimiter = infer_operator_and_type(rule_desc, tuple_ref(factor_ref), tuple_ref(builtin_ref))
        cond = {
            "elementA": placeholder_a,
            "elementAType": "factor",
            "elementADisplay": placeholder_a,
            "operator": operator,
            "dataType": data_type,
            "elementB": placeholder_b,
            "elementBType": "variable",
            "elementBDisplay": placeholder_b,
            "logicToNext": None,
            "stringReplacements": sr,
            "delimiter": delimiter,
            "arrayKeys": None,
        }
        conditions.append(cond)
        groups.append({
            "logicToNext": None,
            "groupFailReason": nopassreason or "",
            "conditions": conditions,
        })

    elif len(refs) == 2 and len(factor_refs) == 2:
        # 双要素比较
        ref_a, ref_b = refs[0], refs[1]
        placeholder_a = placeholder_of(ref_a)
        placeholder_b = placeholder_of(ref_b)

        operator, data_type, sr, delimiter = infer_operator_and_type(rule_desc, tuple_ref(ref_a), tuple_ref(ref_b))

        cond = {
            "elementA": placeholder_a,
            "elementAType": element_type_of(ref_a),
            "elementADisplay": placeholder_a,
            "operator": operator,
            "dataType": data_type,
            "elementB": placeholder_b,
            "elementBType": element_type_of(ref_b),
            "elementBDisplay": placeholder_b,
            "logicToNext": None,
            "stringReplacements": sr,
            "delimiter": delimiter,
            "arrayKeys": None,
        }
        conditions.append(cond)
        groups.append({
            "logicToNext": None,
            "groupFailReason": nopassreason or "",
            "conditions": conditions,
        })

    else:
        if builtin_refs:
            return build_keypoint_rule1(
                kpname,
                rule_desc,
                passreason,
                nopassreason,
                material_name,
                ordernum,
                exclude_situations,
                factor_name=factor_name,
                known_factor_names=known_factor_names,
            )
        # 多要素: 逐对构建多组或多条件
        for i, ref in enumerate(factor_refs):
            placeholder = placeholder_of(ref)
            operator, data_type, sr, delimiter = infer_operator_and_type(rule_desc, tuple_ref(ref))
            is_last = (i == len(refs) - 1)

            if operator in ("notblank", "blank"):
                cond = {
                    "elementA": placeholder,
                    "elementAType": "factor",
                    "elementADisplay": placeholder,
                    "operator": "notblank",
                    "dataType": "string",
                    "elementB": "",
                    "elementBType": "value",
                    "elementBDisplay": "",
                    "logicToNext": None if is_last else "AND",
                    "stringReplacements": None,
                    "delimiter": None,
                    "arrayKeys": None,
                }
            else:
                cond = {
                    "elementA": placeholder,
                    "elementAType": "factor",
                    "elementADisplay": placeholder,
                    "operator": "notblank",
                    "dataType": "string",
                    "elementB": "",
                    "elementBType": "value",
                    "elementBDisplay": "",
                    "logicToNext": None if is_last else "AND",
                    "stringReplacements": None,
                    "delimiter": None,
                    "arrayKeys": None,
                }
            conditions.append(cond)
        group_fail = nopassreason or ""
        groups.append({
            "logicToNext": None,
            "groupFailReason": group_fail,
            "conditions": conditions,
        })

    return {
        "kpname": kpname,
        "content": "",
        "review_rule_text": rule_desc,
        "passreason": passreason or "",
        "nopassreason": nopassreason or "",
        "review_rule": "2",
        "review_conditions": {"groups": groups},
        "is_point": "0",
        "is_contrast": "1",
        "pre_rule_enabled": 0,
        "pre_conditions": None,
        **({"ordernum": ordernum} if ordernum is not None else {}),
        **({"exclude_situations": exclude_situations} if exclude_situations else {}),
    }


def build_keypoint_rule1(kpname, rule_desc, passreason, nopassreason, material_name, ordernum, exclude_situations, factor_name="", known_factor_names=None):
    """构建 review_rule="1" (大模型LLM) 的要点JSON。"""
    # 将 #材料-字段# 和 #内置变量# 转换为统一占位符格式用于LLM提示词
    content = replace_rule_refs_with_placeholders(
        rule_desc or "",
        current_material=material_name,
        known_factor_names=known_factor_names,
    )
    if not content:
        content = f"请审查{material_name}中的{kpname}是否符合要求。"

    return {
        "kpname": kpname,
        "content": content,
        "review_rule_text": rule_desc,
        "passreason": passreason or "",
        "nopassreason": nopassreason or "",
        "review_rule": "1",
        "is_point": "0",
        **({"ordernum": ordernum} if ordernum is not None else {}),
        **({"exclude_situations": exclude_situations} if exclude_situations else {}),
    }


def build_keypoint_rule3(kpname, rule_desc, passreason, nopassreason, material_name, ordernum, exclude_situations, factor_name="", known_factor_names=None):
    """构建 review_rule="3" (Groovy脚本) 的要点JSON。"""
    refs = [
        ref
        for ref in extract_review_refs(
            rule_desc,
            current_material=material_name,
            known_factor_names=known_factor_names,
        )
        if ref["kind"] != "invalid"
    ]
    factor_refs = [ref for ref in refs if ref["kind"] == "factor"]
    builtin_refs = [ref for ref in refs if ref["kind"] == "builtin"]
    if not factor_refs and factor_name:
        factor_refs = [{
            "kind": "factor",
            "token": f"{material_name}-{factor_name}",
            "material": material_name,
            "field": factor_name,
        }]

    # 生成基础Groovy脚本模板
    if factor_refs or builtin_refs:
        input_lines = "\n".join([
            f'def {_sanitize_var(ref["field"])} = input.get("{str(ref["material"]).replace("：", ":").strip()}:{str(ref["field"]).replace("：", ":").strip()}")'
            for ref in factor_refs
        ])
        builtin_lines = "\n".join([
            f'def {_sanitize_var(ref["name"])} = input.get("系统变量:{str(ref["name"]).replace("：", ":").strip()}")'
            for ref in builtin_refs
        ])
        lines = "\n".join(filter(None, [input_lines, builtin_lines]))
        check_names = [
            _sanitize_var(ref["field"])
            for ref in factor_refs
        ] + [
            _sanitize_var(ref["name"])
            for ref in builtin_refs
        ]
        var_checks = " || ".join([f"{name} == null" for name in check_names]) or "false"
        script = (
            f"{lines}\n"
            f"if ({var_checks}) {{\n"
            f'    return [pass: false, reason: "必要字段未识别到"]\n'
            f"}}\n"
            f"// TODO: 在此实现具体的审查逻辑\n"
            f"// 规则: {rule_desc}\n"
            f'return [pass: true, reason: "审查通过"]'
        )
    else:
        script = (
            f"// 规则: {rule_desc}\n"
            f"// TODO: 实现审查逻辑\n"
            f'return [pass: true, reason: "审查通过"]'
        )

    return {
        "kpname": kpname,
        "content": "",
        "review_rule_text": rule_desc,
        "passreason": passreason or "",
        "nopassreason": nopassreason or "",
        "review_rule": "3",
        "review_rule_js": script,
        **({"ordernum": ordernum} if ordernum is not None else {}),
        **({"exclude_situations": exclude_situations} if exclude_situations else {}),
    }


def _sanitize_var(name):
    """将中文字段名转为合法的变量名（拼音首字母简化处理）。"""
    # 简单处理：去除特殊字符，用下划线替代空格
    result = re.sub(r'[^\w\u4e00-\u9fff]', '_', name)
    return f"var_{abs(hash(name)) % 10000}"


# ─────────────────────────── 数据库查库 ─────────────────────────────

def lookup_review_rules_from_db(keypoints_info, item_name, material_name):
    """从审查规则库中按 (item_name, materialname, kpname) 精确查找已导入的规则。

    通过环境变量 AUTO_PROMPT_API_URL 获取后端地址，调用 /api/review-rules/lookup。
    返回 dict: {kpname: rule_dict}，未命中的不在其中。
    """
    api_url = os.environ.get("AUTO_PROMPT_API_URL", "").rstrip("/")
    api_token = os.environ.get("AUTO_PROMPT_API_TOKEN", "")
    if not api_url:
        return {}

    import urllib.request
    import urllib.parse

    matched = {}
    headers = {}
    if api_token:
        headers["Authorization"] = f"Bearer {api_token}"

    for kp in keypoints_info:
        kpname = kp["kpname"]
        params = urllib.parse.urlencode({
            "item_name": item_name,
            "materialname": material_name,
            "kpname": kpname,
        })
        url = f"{api_url}/api/review-rules/lookup?{params}"
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                if data.get("found") and data.get("rule"):
                    matched[kpname] = data["rule"]
                    print(f"  [审查规则库] ✓ '{kpname}' 命中库中已有规则")
        except Exception:
            pass  # 静默失败，回退到 LLM/本地推断

    return matched


# ─────────────────────────── 主处理流程 ─────────────────────────────

def process_material_rules(material_name, keypoints_info, use_llm=False,
                           api_key=None, base_url=None, model=None, timeout=120):
    """为单个材料处理所有审查要点，生成 keypoints 数组。"""
    import time
    keypoints = []
    total_count = len(keypoints_info)
    material_factor_names = {
        str(kp.get("factor_name") or "").strip()
        for kp in keypoints_info
        if str(kp.get("factor_name") or "").strip()
    }

    for idx, kp in enumerate(keypoints_info):
        kp_start_time = time.time()
        kpname = kp["kpname"]
        rule_desc = kp["rule_desc"]
        factor_name = kp.get("factor_name", "")
        passreason = kp["passreason"]
        nopassreason = kp["nopassreason"]
        ordernum = kp["ordernum"]
        exclude_situations = kp["exclude_situations"]
        special_note = kp["special_note"]

        # 输出进度: (当前索引+1)/总数
        print(f"  [进度] {idx + 1}/{total_count} - 要点: {kpname}")
        print(f"         规则说明: {rule_desc[:60]}{'...' if len(rule_desc) > 60 else ''}")

        # 使用LLM推理（如果启用且有API key）
        if use_llm and api_key and rule_desc:
            llm_result = call_llm_for_rule(
                kpname, rule_desc, material_name, api_key, base_url, model, timeout
            )
            if llm_result:
                review_rule = llm_result.get("review_rule", "1")
                kp_json = {
                    "kpname": kpname,
                    "content": llm_result.get("content", ""),
                    "review_rule_text": llm_result.get("review_rule_text", rule_desc),
                    # 规则：优先使用 Excel 列值；Excel 为空则保持空
                    "passreason": passreason or "",
                    "nopassreason": nopassreason or "",
                    "review_rule": review_rule,
                }
                if review_rule == "2" and llm_result.get("review_conditions"):
                    kp_json["review_conditions"] = llm_result["review_conditions"]
                    kp_json["is_contrast"] = "1"
                elif review_rule == "3" and llm_result.get("review_rule_js"):
                    kp_json["review_rule_js"] = llm_result["review_rule_js"]
                if ordernum is not None:
                    kp_json["ordernum"] = ordernum
                if exclude_situations:
                    kp_json["exclude_situations"] = exclude_situations
                print(f"         [LLM] -> review_rule={review_rule}")
                keypoints.append(kp_json)
                continue

        # 本地推断
        review_rule = infer_review_rule_local(
            rule_desc,
            special_note,
            factor_name=factor_name,
            material_name=material_name,
            known_factor_names=material_factor_names,
        )
        print(f"         [本地推断] -> review_rule={review_rule}")

        if review_rule == "2":
            kp_json = build_keypoint_rule2(
                kpname, rule_desc, passreason, nopassreason,
                material_name, ordernum, exclude_situations,
                factor_name=factor_name,
                known_factor_names=material_factor_names,
            )
        elif review_rule == "3":
            kp_json = build_keypoint_rule3(
                kpname, rule_desc, passreason, nopassreason,
                material_name, ordernum, exclude_situations,
                factor_name=factor_name,
                known_factor_names=material_factor_names,
            )
        else:
            kp_json = build_keypoint_rule1(
                kpname, rule_desc, passreason, nopassreason,
                material_name, ordernum, exclude_situations,
                factor_name=factor_name,
                known_factor_names=material_factor_names,
            )

        keypoints.append(kp_json)

        # 输出耗时
        kp_elapsed = time.time() - kp_start_time
        print(f"         [完成] 耗时: {kp_elapsed:.2f}秒")

    return keypoints


# ─────────────────────────── 入口 ─────────────────────────────────

def main(work_dir, use_llm=False, api_key=None, base_url=None, model=None, timeout=120, material_filter=None):
    work_dir = os.path.abspath(work_dir)

    if not os.path.isdir(work_dir):
        print(f"[错误] 工作目录不存在: {work_dir}")
        sys.exit(1)

    # 查找 factors.xlsx
    factors_path = None
    for fname in ['factors.xlsx']:
        p = os.path.join(work_dir, fname)
        if os.path.exists(p):
            factors_path = p
            break

    if not factors_path:
        print("[错误] 未找到 factors.xlsx 文件")
        sys.exit(1)

    print(f"[信息] 读取审查要点定义: {factors_path}")
    rules_dict = read_review_rules_from_excel(factors_path)

    if not rules_dict:
        print("[错误] 未读取到任何审查要点信息（审查要点名称列为空）")
        sys.exit(1)

    # 按用户选择过滤材料
    if material_filter:
        filter_set = set(material_filter)
        rules_dict = {k: v for k, v in rules_dict.items() if k in filter_set}
        if not rules_dict:
            print(f"[错误] 指定的材料在 Excel 中均未找到: {material_filter}")
            sys.exit(1)

    import time
    total_kp = sum(len(v) for v in rules_dict.values())
    print(f"[信息] 共读取 {len(rules_dict)} 个材料，{total_kp} 个审查要点\n")

    # 获取事项名称（从上级目录名或环境变量）
    item_name = os.environ.get("ITEM_NAME", "")
    if not item_name:
        item_name = os.path.basename(os.path.dirname(work_dir)) if os.path.basename(work_dir) != work_dir else ""

    results = []
    material_start_time = time.time()
    for idx, (material_name, keypoints_info) in enumerate(rules_dict.items()):
        mat_start_time = time.time()
        # 输出材料进度: (当前索引+1)/总数
        print(f"\n[进度] 材料 {idx + 1}/{len(rules_dict)} - {material_name}")

        try:
            # 优先从审查规则库精确匹配
            db_matched = lookup_review_rules_from_db(keypoints_info, item_name, material_name)

            db_keypoints = []
            remaining_kps = []
            for kp in keypoints_info:
                kpname = kp["kpname"]
                if kpname in db_matched:
                    rule = db_matched[kpname]
                    # 构造与 process_material_rules 输出相同的格式
                    kp_json = {
                        "kpname": kpname,
                        "content": rule.get("content", ""),
                        "review_rule_text": rule.get("review_rule_text", ""),
                        "passreason": rule.get("passreason", ""),
                        "nopassreason": rule.get("nopassreason", ""),
                        "review_rule": rule.get("review_rule", "1"),
                    }
                    if rule.get("review_conditions"):
                        kp_json["review_conditions"] = rule["review_conditions"]
                        kp_json["is_contrast"] = "1"
                    if rule.get("review_rule_js"):
                        kp_json["review_rule_js"] = rule["review_rule_js"]
                    db_keypoints.append(kp_json)
                else:
                    remaining_kps.append(kp)

            if db_matched:
                print(f"  [审查规则库] 命中 {len(db_matched)} 个，剩余 {len(remaining_kps)} 个需生成")

            # 对未命中的要素走原有 LLM/本地推断流程
            generated_keypoints = []
            if remaining_kps:
                generated_keypoints = process_material_rules(
                    material_name, remaining_kps,
                    use_llm=use_llm, api_key=api_key, base_url=base_url, model=model, timeout=timeout
                )

            keypoints = db_keypoints + generated_keypoints

            import_json = {
                "materialname": material_name,
                "keypoints": keypoints,
            }

            # 输出到材料子目录
            material_dir = os.path.join(work_dir, material_name)
            if os.path.isdir(material_dir):
                output_path = os.path.join(material_dir, f"{material_name}--审查规则导入.json")
            else:
                output_path = os.path.join(work_dir, f"{material_name}--审查规则导入.json")

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(import_json, f, ensure_ascii=False, indent=2)

            print(f"  [完成] 生成: {output_path}")
            print(f"         审查要点: {len(keypoints)} 个")

            # 输出材料耗时
            mat_elapsed = time.time() - mat_start_time
            print(f"         耗时: {mat_elapsed:.2f}秒\n")

            results.append({
                "material": material_name,
                "success": True,
                "output": output_path,
                "keypoint_count": len(keypoints),
            })

        except Exception as e:
            import traceback
            print(f"  [错误] 生成失败: {e}\n")
            traceback.print_exc()
            results.append({
                "material": material_name,
                "success": False,
                "error": str(e),
                "keypoint_count": 0,
            })

    ok = sum(1 for r in results if r["success"])
    total_elapsed = time.time() - material_start_time
    print(f"[完成] 所有材料处理完毕：成功 {ok} / {len(results)} 个，总耗时: {total_elapsed:.2f}秒")

    # 输出结构化结果供桌面端解析
    print("RESULTS_JSON:" + json.dumps(results, ensure_ascii=False))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='审查规则JSON生成器')
    parser.add_argument('work_dir', help='材料集目录路径（含 factors.xlsx）')
    parser.add_argument('--use-llm', action='store_true', help='启用LLM推理（需配置API Key）')
    parser.add_argument('--api-key', default='', help='LLM API Key')
    parser.add_argument('--base-url', default='https://api.openai.com/v1', help='LLM API Base URL')
    parser.add_argument('--model', default='gpt-4o-mini', help='LLM 模型名称')
    parser.add_argument('--timeout', type=int, default=120, help='LLM 调用超时时间（秒）')
    parser.add_argument('--materials', nargs='*', default=None,
                        help='仅处理指定的材料名称（不指定则处理全部）')
    args = parser.parse_args()

    main(
        args.work_dir,
        use_llm=args.use_llm,
        api_key=args.api_key,
        base_url=args.base_url,
        model=args.model,
        timeout=args.timeout,
        material_filter=args.materials,
    )
