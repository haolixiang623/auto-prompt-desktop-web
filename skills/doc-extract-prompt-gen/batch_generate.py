#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量提示词生成器
从父目录的 factors 文件（CSV 或 Excel）中读取所有材料名称，为每个材料目录生成提示词
"""

import os
import sys
import csv
import subprocess

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

def get_materials_from_file(file_path):
    """从要素文件（CSV 或 Excel）中提取所有唯一的材料名称"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"未找到文件: {file_path}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext in ['.xlsx', '.xls']:
        return _get_materials_from_excel(file_path)
    elif file_ext == '.csv':
        return _get_materials_from_csv(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {file_ext}")

def _get_materials_from_csv(csv_path):
    """从 CSV 文件中提取材料名称"""
    materials = set()
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        if '材料名称' not in reader.fieldnames:
            raise ValueError(f"CSV必须包含'材料名称'列，当前列: {reader.fieldnames}")
        
        for row in reader:
            material = row.get('材料名称', '').strip()
            if material:
                materials.add(material)
    
    return sorted(materials)

def _get_materials_from_excel(excel_path):
    """从 Excel 文件中提取材料名称"""
    if not EXCEL_SUPPORT:
        raise ImportError("Excel 支持需要安装 openpyxl。安装命令: pip install openpyxl")
    
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 读取表头
    headers = [cell.value for cell in ws[1]]
    
    if '材料名称' not in headers:
        raise ValueError(f"Excel必须包含'材料名称'列，当前列: {headers}")
    
    material_col_idx = headers.index('材料名称')
    
    materials = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > material_col_idx:
            material = str(row[material_col_idx]).strip() if row[material_col_idx] else ''
            if material:
                materials.add(material)
    
    wb.close()
    return sorted(materials)

def main():
    # 获取父目录路径
    if len(sys.argv) > 1:
        parent_dir = os.path.abspath(sys.argv[1])
    else:
        parent_dir = os.getcwd()
    
    if not os.path.isdir(parent_dir):
        print(f"[错误] 指定的目录不存在: {parent_dir}")
        return
    
    # 查找要素文件（优先 Excel 格式）
    def find_factors_file(base_dir):
        """查找要素文件，优先 Excel 格式"""
        for ext in ['.xlsx', '.xls', '.csv']:
            file_path = os.path.join(base_dir, f'factors{ext}')
            if os.path.exists(file_path):
                return file_path
        return None
    
    factors_file = find_factors_file(parent_dir)
    
    if not factors_file:
        print(f"[错误] 未找到要素文件 (factors.csv/factors.xlsx/factors.xls)")
        return
    
    try:
        print("="*60)
        print("批量文档要素提取提示词生成器")
        print("="*60)
        print(f"\n[父目录] {parent_dir}")
        print(f"[文件] {factors_file}")
        
        # 1. 从文件中提取材料名称
        materials = get_materials_from_file(factors_file)
        print(f"\n✓ 发现 {len(materials)} 个材料: {', '.join(materials)}")
        
        # 2. 获取 generate_prompt.py 脚本路径
        script_dir = os.path.dirname(os.path.abspath(__file__))
        generate_script = os.path.join(script_dir, 'generate_prompt.py')
        
        if not os.path.exists(generate_script):
            print(f"\n[错误] 未找到生成脚本: {generate_script}")
            return
        
        # 3. 为每个材料生成提示词
        success_count = 0
        failed_materials = []
        
        for i, material in enumerate(materials, 1):
            print(f"\n{'='*60}")
            print(f"[{i}/{len(materials)}] 处理材料: {material}")
            print(f"{'='*60}")
            
            # 检查材料目录是否存在
            material_dir = os.path.join(parent_dir, material)
            if not os.path.isdir(material_dir):
                print(f"[警告] 材料目录不存在: {material_dir}")
                failed_materials.append(material)
                continue
            
            # 调用 generate_prompt.py
            try:
                cmd = [
                    sys.executable,
                    generate_script,
                    material_dir,
                    material
                ]
                
                result = subprocess.run(
                    cmd,
                    capture_output=False,
                    text=True,
                    env={**os.environ}
                )
                
                if result.returncode == 0:
                    success_count += 1
                    print(f"\n✓ {material} 处理成功")
                else:
                    failed_materials.append(material)
                    print(f"\n✗ {material} 处理失败")
            
            except Exception as e:
                print(f"\n[错误] {material} 处理异常: {e}")
                failed_materials.append(material)
        
        # 4. 输出统计结果
        print(f"\n{'='*60}")
        print("批量处理完成")
        print(f"{'='*60}")
        print(f"✓ 成功: {success_count}/{len(materials)}")
        
        if failed_materials:
            print(f"✗ 失败: {len(failed_materials)}")
            print(f"  失败材料: {', '.join(failed_materials)}")
        
    except Exception as e:
        print(f"\n[错误] 批量处理失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
