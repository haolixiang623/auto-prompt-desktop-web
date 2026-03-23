import os
import csv
import base64
import json
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("[警告] 未安装 openpyxl，Excel 支持不可用。安装命令: pip install openpyxl")


# DashScope 通过 OpenAI 兼容接口传递时，这些参数必须放在 extra_body 而非顶层 kwargs
_DASHSCOPE_BODY_PARAMS = {
    'enable_thinking', 'thinking_budget', 'translation_options',
    'vl_high_resolution_images', 'search_options',
}

def get_extra_params():
    """读取桌面端传入的额外模型参数，拆分为标准参数和 DashScope 专有参数(extra_body)"""
    raw = os.environ.get("GENERATE_EXTRA_PARAMS", "{}")
    try:
        all_params = json.loads(raw)
    except Exception:
        all_params = {}
    standard = {k: v for k, v in all_params.items() if k not in _DASHSCOPE_BODY_PARAMS}
    body = {k: v for k, v in all_params.items() if k in _DASHSCOPE_BODY_PARAMS}
    if body:
        standard['extra_body'] = body
    return standard

def load_template(template_path):
    """加载模板文件，优先使用工作目录的模板，否则使用 skill 目录的默认模板"""
    if os.path.exists(template_path):
        print(f"[模板] 使用工作目录模板: {template_path}")
        with open(template_path, 'r', encoding='utf-8') as f:
            return f.read()
    
    # 如果工作目录没有模板，使用 skill 目录的默认模板
    skill_dir = get_skill_dir()
    default_template_path = os.path.join(skill_dir, 'template.txt')
    
    if os.path.exists(default_template_path):
        print(f"[模板] 使用默认模板: {default_template_path}")
        with open(default_template_path, 'r', encoding='utf-8') as f:
            return f.read()
    
    raise FileNotFoundError(f"未找到模板文件: {template_path} 或默认模板: {default_template_path}")

def parse_factors_file(file_path, material_name=None):
    """解析要素文件（支持 CSV 和 Excel），支持要素名称、提取说明、规则说明三列，可选材料名称过滤
    
    Args:
        file_path: CSV 或 Excel 文件路径
        material_name: 可选，材料名称，用于过滤特定材料的要素
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"未找到文件: {file_path}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    # 根据文件扩展名选择解析方式
    if file_ext in ['.xlsx', '.xls']:
        return _parse_excel(file_path, material_name)
    elif file_ext == '.csv':
        return _parse_csv(file_path, material_name)
    else:
        raise ValueError(f"不支持的文件格式: {file_ext}，仅支持 .csv, .xlsx, .xls")

def _parse_csv(csv_path, material_name=None):
    """解析 CSV 文件"""
    factors = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        if '要素名称' not in reader.fieldnames:
            raise ValueError(f"CSV必须包含'要素名称'列，当前列: {reader.fieldnames}")
        
        # 检查是否有材料名称列
        has_material_column = '材料名称' in reader.fieldnames
        
        for row in reader:
            name = row.get('要素名称', '').strip()
            if not name:  # 跳过空行
                continue
            
            # 如果指定了材料名称且CSV有材料名称列，则过滤
            if material_name and has_material_column:
                row_material = row.get('材料名称', '').strip()
                if row_material != material_name:
                    continue
            
            factor = {
                'name': name,
                'extract_desc': row.get('要素提取说明', '').strip(),
                'rule_desc': row.get('提取规则说明', '').strip()
            }
            factors.append(factor)
    
    return factors

def _parse_excel(excel_path, material_name=None):
    """解析 Excel 文件，自动适配两种格式：
    - 简单格式：含「要素名称」列头
    - 扩展格式：A=事项名称, B=材料名称, D=要素字段名称, G=要素提取说明（跨行合并）
    """
    if not EXCEL_SUPPORT:
        raise ImportError("Excel 支持需要安装 openpyxl。安装命令: pip install openpyxl")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取表头
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

    # 自动检测格式
    is_extended = (
        ('事项' in headers[0]) or
        (len(headers) > 1 and '材料名称' in headers[1])
    )

    factors = []

    if is_extended:
        # 扩展格式：B=材料名称(col1), D=要素字段名称(col3), G=要素提取说明(col6)
        print(f"[Excel] 检测到扩展格式（B=材料名称, D=要素字段名称）")
        current_material = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_mat = row[1] if len(row) > 1 else None
            if raw_mat:
                s = str(raw_mat).strip()
                if s and '\n' not in s and len(s) < 60:
                    current_material = s
            if not current_material:
                continue
            # 如果指定了材料名称，只处理匹配的材料
            if material_name and current_material != material_name:
                continue
            factor_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            factor_usage = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            if not factor_name or '\n' in factor_name or len(factor_name) > 50:
                continue
            factors.append({
                'name': factor_name,
                'extract_desc': factor_usage,
                'rule_desc': ''
            })
    else:
        # 简单格式：含「要素名称」列
        if '要素名称' not in headers:
            raise ValueError(f"Excel必须包含'要素名称'列，当前列: {headers}")
        col_map = {h: i for i, h in enumerate(headers)}
        has_material_column = '材料名称' in headers
        current_material = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # 继承材料名称（合并单元格）
            if has_material_column:
                raw_mat = row[col_map['材料名称']] if col_map['材料名称'] < len(row) else None
                if raw_mat:
                    s = str(raw_mat).strip()
                    if s:
                        current_material = s
            name_idx = col_map['要素名称']
            if name_idx >= len(row):
                continue
            name = str(row[name_idx]).strip() if row[name_idx] else ''
            if not name:
                continue
            if material_name and has_material_column:
                if current_material != material_name:
                    continue
            factor = {
                'name': name,
                'extract_desc': str(row[col_map['要素提取说明']]).strip() if col_map.get('要素提取说明', -1) >= 0 and col_map['要素提取说明'] < len(row) and row[col_map['要素提取说明']] else '',
                'rule_desc': str(row[col_map['提取规则说明']]).strip() if col_map.get('提取规则说明', -1) >= 0 and col_map['提取规则说明'] < len(row) and row[col_map['提取规则说明']] else '',
            }
            factors.append(factor)

    wb.close()
    return factors

def get_skill_dir():
    """获取 skill 基础目录"""
    # 查找 .windsurf/skills/doc-extract-prompt-gen 目录
    current = os.path.dirname(os.path.abspath(__file__))
    
    # 如果当前就在 skill 目录下
    if 'doc-extract-prompt-gen' in current and '.windsurf' in current:
        return current
    
    # 优先查找 .windsurf/skills 目录
    home = os.path.expanduser('~')
    windsurf_skill = os.path.join(home, 'Desktop', 'projects', 'Auto-Prompt', '.windsurf', 'skills', 'doc-extract-prompt-gen')
    if os.path.exists(windsurf_skill):
        return windsurf_skill
    
    # 备选：用户主目录下的 skill 目录
    skill_dir = os.path.join(home, '.claude', 'skills', 'doc-extract-prompt-gen')
    if os.path.exists(skill_dir):
        return skill_dir
    
    # 如果都找不到，使用当前脚本目录
    return current

def load_case_library():
    """加载案例库（从 skill 目录）"""
    skill_dir = get_skill_dir()
    case_lib_path = os.path.join(skill_dir, 'case_library.json')
    
    print(f"[案例库] 从 {case_lib_path} 加载")
    
    if not os.path.exists(case_lib_path):
        print("\n[提示] 未找到案例库文件，将创建新的案例库")
        return {"version": "1.0", "cases": []}
    
    try:
        with open(case_lib_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"\n[警告] 案例库加载失败: {e}，将使用空案例库")
        return {"version": "1.0", "cases": []}

def save_case_library(case_lib):
    """保存案例库（到 skill 目录）"""
    skill_dir = get_skill_dir()
    case_lib_path = os.path.join(skill_dir, 'case_library.json')
    
    try:
        with open(case_lib_path, 'w', encoding='utf-8') as f:
            json.dump(case_lib, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"\n[错误] 案例库保存失败: {e}")
        return False

def match_cases_with_qwen(client, factors, case_lib, material_name=None):
    """使用 Qwen 智能匹配案例库中的相似案例
    
    Args:
        client: Qwen API 客户端
        factors: 要素列表
        case_lib: 案例库
        material_name: 材料名称，用于过滤案例（可选）
    """
    print("\n[案例库] 正在分析要素并匹配案例库...")
    
    if not case_lib.get('cases'):
        print("[案例库] 案例库为空，跳过匹配")
        return {}
    
    # 根据材料名称过滤案例
    filtered_cases = case_lib['cases']
    if material_name:
        filtered_cases = [c for c in case_lib['cases'] if c.get('material_name') == material_name]
        print(f"[案例库] 按材料名称 '{material_name}' 过滤，找到 {len(filtered_cases)} 个相关案例")
        
        # 如果没有匹配的案例，尝试使用通用案例
        if not filtered_cases:
            filtered_cases = [c for c in case_lib['cases'] if c.get('material_name') == '通用']
            if filtered_cases:
                print(f"[案例库] 未找到专属案例，使用 {len(filtered_cases)} 个通用案例")
            else:
                print("[案例库] 未找到匹配的案例，跳过匹配")
                return {}
    
    # 构建要素描述
    factor_descriptions = []
    for factor in factors:
        desc = f"- {factor['name']}"
        if factor['extract_desc']:
            desc += f": {factor['extract_desc']}"
        if factor['rule_desc']:
            desc += f" ({factor['rule_desc']})"
        factor_descriptions.append(desc)
    
    # 构建案例库描述（包含提取规则和材料名称）
    case_descriptions = []
    for i, case in enumerate(filtered_cases):
        case_desc = f"{i}. [{case.get('material_name', '通用')}] {case['factor_name']}\n"
        case_desc += f"   提取规则: {case.get('extraction_rule', '')[:100]}..."  # 显示前100字符
        if case.get('extract_desc'):
            case_desc += f"\n   提取说明: {case['extract_desc']}"
        if case.get('rule_desc'):
            case_desc += f"\n   规则说明: {case['rule_desc']}"
        case_descriptions.append(case_desc)
    
    match_prompt = f"""请分析以下要素列表，并从案例库中找出最匹配的案例。

当前要素列表：
{chr(10).join(factor_descriptions)}

案例库（包含提取规则）：
{chr(10).join(case_descriptions)}

请为每个要素找出最匹配的案例编号（如果有的话）。匹配标准（按优先级排序）：
1. **要素名称**相同或高度相似（语义匹配）
2. **提取规则**语义相近或描述同一类提取任务
3. 提取说明和规则说明语义相近
4. 如果要素名称和提取规则都不相似，返回 -1

**重要**：只有当要素名称或提取规则高度相似时才匹配，避免错误匹配。

请以JSON格式返回：
{{
  "matches": [
    {{"factor_name": "要素名称", "case_index": 0, "confidence": "high/medium/low"}}
  ]
}}"""
    
    extra = get_extra_params()
    try:
        response = client.chat.completions.create(
            model="qwen3.5-35b-a3b",
            messages=[{"role": "user", "content": match_prompt}],
            **extra
        )
        result = response.choices[0].message.content
        
        # 提取JSON
        if "```json" in result:
            result = result.split("```json")[1].split("```")[0].strip()
        elif "```" in result:
            result = result.split("```")[1].split("```")[0].strip()
        
        matches_data = json.loads(result)
        
        # 构建匹配字典（只接受高置信度匹配）
        matched_cases = {}
        for match in matches_data.get('matches', []):
            case_idx = match.get('case_index', -1)
            confidence = match.get('confidence', 'low')
            
            if case_idx >= 0 and case_idx < len(filtered_cases):
                factor_name = match.get('factor_name')
                case = filtered_cases[case_idx]
                
                # 只接受高置信度或中等置信度的匹配
                if confidence in ['high', 'medium']:
                    matched_cases[factor_name] = case
                    material_tag = f"[{case.get('material_name', '通用')}]"
                    print(f"[案例库] ✓ '{factor_name}' 匹配到 {material_tag} '{case['factor_name']}' (置信度: {confidence})")
                else:
                    print(f"[案例库] ✗ '{factor_name}' 匹配置信度过低 ({confidence})，将使用AI生成")
        
        return matched_cases
    except Exception as e:
        print(f"\n[警告] 案例库匹配失败: {e}，将使用AI生成")
        return {}

def add_cases_to_library(case_lib, factors_with_rules, material_name=None, source="ai_generated"):
    """将验证通过的规则添加到案例库
    
    Args:
        case_lib: 案例库
        factors_with_rules: 要素规则列表
        material_name: 材料名称（可选）
        source: 来源标记
    """
    print("\n[案例库] 正在保存新规则到案例库...")
    
    added_count = 0
    for factor in factors_with_rules:
        # 检查是否已存在完全相同的案例（基于材料名称+要素名称）
        exists = False
        for case in case_lib['cases']:
            if (case.get('material_name') == material_name and
                case['factor_name'] == factor['name'] and 
                case.get('extract_desc') == factor.get('extract_desc', '') and
                case.get('rule_desc') == factor.get('rule_desc', '')):
                exists = True
                break
        
        if not exists:
            new_case = {
                "material_name": material_name or "通用",
                "factor_name": factor['name'],
                "extract_desc": factor.get('extract_desc', ''),
                "rule_desc": factor.get('rule_desc', ''),
                "extraction_rule": factor['rule'],
                "format_requirement": factor['format'],
                "source": source,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "tags": []
            }
            case_lib['cases'].append(new_case)
            added_count += 1
            material_tag = f"[{material_name}]" if material_name else "[通用]"
            print(f"[案例库] ✓ 已添加: {material_tag} {factor['name']} (来源: {source})")
    
    if added_count > 0:
        if save_case_library(case_lib):
            print(f"[案例库] ✓ 成功保存 {added_count} 个新案例到案例库")
        else:
            print(f"[案例库] ✗ 保存失败")
    else:
        print("[案例库] 所有规则已存在于案例库中")

def generate_factors_text(factors):
    lines = []
    for f in factors:
        lines.append(f"## {f['index']}.{f['name']}")
        
        rule_text = f['rule']
        format_text = f['format']
        
        if format_text:
            lines.append(f"{rule_text}，{format_text}")
        else:
            lines.append(f"{rule_text}")
    
    return "\n".join(lines)

def build_prompt(template_text, factors_text):
    if '$(factors)' not in template_text:
        print("警告: 模板中未找到 '$(factors)' 占位符。")
    return template_text.replace('$(factors)', factors_text)

def pdf_to_images(pdf_path, output_dir):
    """将 PDF 每页转为 PNG 图片，返回图片路径列表。
    优先使用 PyMuPDF (fitz)，其次 pdf2image，都没有则返回空列表。"""
    base_name = Path(pdf_path).stem
    pages = []
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(pdf_path)
        for i, page in enumerate(doc):
            mat = fitz.Matrix(2.0, 2.0)  # 2x scale for clarity
            pix = page.get_pixmap(matrix=mat)
            out_path = os.path.join(output_dir, f"{base_name}_p{i+1}.png")
            pix.save(out_path)
            pages.append(out_path)
        doc.close()
        print(f"[PDF] 使用 PyMuPDF 将 {Path(pdf_path).name} 转换为 {len(pages)} 张图片")
        return pages
    except ImportError:
        pass
    try:
        from pdf2image import convert_from_path
        imgs = convert_from_path(pdf_path, dpi=200)
        for i, img in enumerate(imgs):
            out_path = os.path.join(output_dir, f"{base_name}_p{i+1}.png")
            img.save(out_path, "PNG")
            pages.append(out_path)
        print(f"[PDF] 使用 pdf2image 将 {Path(pdf_path).name} 转换为 {len(pages)} 张图片")
        return pages
    except ImportError:
        pass
    print(f"[警告] 无法转换 PDF 文件 {Path(pdf_path).name}，")
    print("  请安装 PyMuPDF: pip install pymupdf  或  pdf2image: pip install pdf2image")
    return []


def get_images_in_dir(directory):
    """获取目录中的图片文件；PDF 文件自动转换为图片。"""
    valid_img_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.webp', '.gif'}
    images = []
    pdf_tmp_dir = os.path.join(directory, '__pdf_converted__')

    for file in sorted(os.listdir(directory)):
        ext = Path(file).suffix.lower()
        full_path = os.path.join(directory, file)
        if ext in valid_img_exts:
            images.append(full_path)
        elif ext == '.pdf':
            os.makedirs(pdf_tmp_dir, exist_ok=True)
            converted = pdf_to_images(full_path, pdf_tmp_dir)
            images.extend(converted)

    return images

def get_qwen_client():
    """获取 Qwen API 客户端"""
    try:
        from openai import OpenAI
    except ImportError:
        print("\n[错误] 缺少 openai 库，请执行: pip install openai")
        return None

    api_key = os.environ.get("DASHSCOPE_API_KEY")
    if not api_key:
        print("\n[错误] 未检测到 DASHSCOPE_API_KEY 环境变量")
        print("请设置环境变量: export DASHSCOPE_API_KEY='your-api-key'")
        return None
    
    return OpenAI(
        api_key=api_key,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    )

def analyze_image_with_qwen(client, factors, image_paths):
    """让 Qwen 分析所有图片并识别要素（利用CSV中的上下文信息），合并各图片的识别结果"""
    if isinstance(image_paths, str):
        image_paths = [image_paths]
    
    # 构建带上下文的要素列表
    factor_list = []
    for i, factor in enumerate(factors, 1):
        factor_desc = f"{i}. {factor['name']}"
        if factor['extract_desc']:
            factor_desc += f" - {factor['extract_desc']}"
        if factor['rule_desc']:
            factor_desc += f"（{factor['rule_desc']}）"
        factor_list.append(factor_desc)
    
    analysis_prompt = f"""请仔细分析这张图片，识别以下要素的内容：
{chr(10).join(factor_list)}

对于每个要素，请提取其在图片中的实际值。如果某个要素不存在，请说明。
请以JSON格式返回，格式如下：
{{
  "factors": [
    {{"name": "要素名称", "value": "识别到的值", "exists": true}}
  ]
}}"""
    
    # 逐张分析，合并结果（已找到的要素不再被后续图片覆盖为不存在）
    merged = {}  # factor_name -> result_dict
    
    for image_path in image_paths:
        print(f"\n[步骤1] 正在分析图片: {os.path.basename(image_path)}")
        print("[步骤1] 调用 Qwen 识别要素...")
        
        with open(image_path, "rb") as image_file:
            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
        
        extra = get_extra_params()
        try:
            response = client.chat.completions.create(
                model="qwen-vl-max",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": analysis_prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                **extra
            )
            result_str = response.choices[0].message.content
            print(f"[步骤1] 识别结果:")
            print(result_str)
            
            # 解析并合并：已有值的要素不被 exists=false 的结果覆盖
            parsed = _parse_analysis_result(result_str)
            for item in parsed:
                name = item.get('name', '')
                if not name:
                    continue
                existing = merged.get(name)
                if existing is None or (not existing.get('exists') and item.get('exists')):
                    merged[name] = item
        except Exception as e:
            print(f"\n[错误] 调用 Qwen API 失败 ({os.path.basename(image_path)}): {e}")
    
    if not merged:
        return None
    
    # 重新序列化为 JSON 字符串供后续步骤使用
    combined = json.dumps({"factors": list(merged.values())}, ensure_ascii=False, indent=2)
    print(f"\n[步骤1] 合并后的识别结果（共 {len(image_paths)} 张图片）:")
    print(combined)
    return combined


def _parse_analysis_result(result_str):
    """从 Qwen 返回的字符串中提取 factors 列表"""
    try:
        if "```json" in result_str:
            result_str = result_str.split("```json")[1].split("```")[0].strip()
        elif "```" in result_str:
            result_str = result_str.split("```")[1].split("```")[0].strip()
        return json.loads(result_str).get('factors', [])
    except Exception:
        return []

def generate_smart_rules(client, factors, analysis_result):
    """让 Qwen 根据识别结果和CSV上下文生成智能提取规则（具备泛化能力）"""
    print("\n[步骤2] 正在生成智能提取规则...")

    # 从环境变量读取造物主提示词（要素提取专用）
    god_prompt = os.environ.get("EXTRACT_GOD_PROMPT", "").strip()

    # 构建带上下文的要素列表
    factor_context = []
    for i, factor in enumerate(factors, 1):
        context = f"{i}. {factor['name']}"
        if factor['extract_desc']:
            context += f"\n   提取说明: {factor['extract_desc']}"
        if factor['rule_desc']:
            context += f"\n   规则说明: {factor['rule_desc']}"
        factor_context.append(context)

    rule_prompt = f"""基于以下要素识别结果和用户提供的说明，为每个要素生成精准的提取规则和格式要求。

要素列表及说明：
{chr(10).join(factor_context)}

识别结果（仅用于理解要素的位置和上下文，不得将具体数值写入规则）：
{analysis_result}

请为每个要素生成：
1. **提取规则**：描述如何在同类文档中定位和识别该要素（结合用户提供的提取说明和规则说明）
2. **格式要求**：描述提取后的格式处理要求（参考用户的规则说明）

**泛化要求（重要）**：
- 提取规则必须具备通用性，适用于同类型的所有文档，而不仅针对本次识别到的具体文档
- 禁止在规则中出现具体的数值、金额、日期、名称、编号等特定文档才有的内容
- 规则描述应基于要素的结构特征和位置规律（如"位于文档抬头"、"表格第X列"、"盖章处下方"等）
- 可描述数据类型特征（如"数字"、"日期格式"、"中文名称"），但不能写具体值
- 格式要求要明确，参考用户的规则说明，如果不需要格式处理则说明"保持原格式"

请以JSON格式返回：
{{
  "factors": [
    {{
      "name": "要素名称",
      "rule": "通用的提取规则描述（不含具体值）",
      "format": "格式要求描述"
    }}
  ]
}}"""

    # 构建消息：若有造物主提示词则作为 system 消息
    if god_prompt:
        messages = [
            {"role": "system", "content": god_prompt},
            {"role": "user", "content": rule_prompt}
        ]
        print(f"[步骤2] 使用造物主提示词（{len(god_prompt)} 字符）")
    else:
        messages = [{"role": "user", "content": rule_prompt}]

    extra = get_extra_params()
    try:
        response = client.chat.completions.create(
            model="qwen-vl-max",
            messages=messages,
            **extra
        )
        result = response.choices[0].message.content
        print("\n[步骤2] 生成的智能规则:")
        print(result)
        return result
    except Exception as e:
        print(f"\n[错误] 生成规则失败: {e}")
        return None

def parse_smart_rules(rules_json_str):
    """解析 Qwen 返回的智能规则 JSON"""
    try:
        # 提取 JSON 代码块
        if "```json" in rules_json_str:
            start = rules_json_str.find("```json") + 7
            end = rules_json_str.find("```", start)
            rules_json_str = rules_json_str[start:end].strip()
        elif "```" in rules_json_str:
            start = rules_json_str.find("```") + 3
            end = rules_json_str.find("```", start)
            rules_json_str = rules_json_str[start:end].strip()
        
        rules_data = json.loads(rules_json_str)
        factors = []
        for i, item in enumerate(rules_data.get('factors', []), start=1):
            factors.append({
                'index': i,
                'name': item.get('name', ''),
                'rule': item.get('rule', ''),
                'format': item.get('format', '')
            })
        return factors
    except Exception as e:
        print(f"\n[警告] 解析智能规则失败: {e}")
        return None

def validate_final_prompt(client, prompt, image_paths):
    """验证最终生成的提示词，对所有图片逐张验证"""
    if isinstance(image_paths, str):
        image_paths = [image_paths]
    
    print(f"\n[步骤3] 正在验证最终提示词（共 {len(image_paths)} 张图片）...")
    
    any_success = False
    for image_path in image_paths:
        print(f"\n[步骤3] 验证图片: {os.path.basename(image_path)}")
        with open(image_path, "rb") as image_file:
            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
        
        extra = get_extra_params()
        try:
            response = client.chat.completions.create(
                model="qwen-vl-max",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                **extra
            )
            print(f"=== 验证结果 [{os.path.basename(image_path)}] ===")
            print(response.choices[0].message.content)
            print("=========================================\n")
            any_success = True
        except Exception as e:
            print(f"\n[错误] 验证失败 ({os.path.basename(image_path)}): {e}")
    
    return any_success

def main():
    import sys
    
    # 支持通过命令行参数指定工作目录和材料名称
    # 用法: python3 generate_prompt.py <工作目录> [材料名称]
    if len(sys.argv) > 1:
        current_dir = os.path.abspath(sys.argv[1])
        if not os.path.isdir(current_dir):
            print(f"[错误] 指定的目录不存在: {current_dir}")
            return
    else:
        current_dir = os.getcwd()
    
    # 材料名称（可选）
    material_name = sys.argv[2] if len(sys.argv) > 2 else None
    
    dir_name = os.path.basename(current_dir)
    
    # 要素文件查找逻辑（支持 CSV 和 Excel）：
    # 1. 如果指定了材料名称，优先使用父目录的 factors 文件
    # 2. 否则使用当前目录的 factors 文件
    # 3. 优先级：.xlsx > .xls > .csv
    def find_factors_file(base_dir):
        """查找要素文件，优先 Excel 格式"""
        for ext in ['.xlsx', '.xls', '.csv']:
            file_path = os.path.join(base_dir, f'factors{ext}')
            if os.path.exists(file_path):
                return file_path
        return None
    
    if material_name:
        parent_dir = os.path.dirname(current_dir)
        factors_file = find_factors_file(parent_dir) or find_factors_file(current_dir)
    else:
        factors_file = find_factors_file(current_dir)
    
    if not factors_file:
        print(f"[错误] 未找到要素文件 (factors.csv/factors.xlsx/factors.xls)")
        return
    
    template_path = os.path.join(current_dir, 'template.txt')
    output_path = os.path.join(current_dir, f"{dir_name}--要素提取完整提示词.txt")
    
    try:
        print("="*60)
        print("智能文档要素提取提示词生成器")
        print("="*60)
        
        if material_name:
            print(f"\n[材料] 当前材料: {material_name}")
        print(f"[文件] 使用文件: {factors_file}")
        
        # 1. 解析要素文件（CSV 或 Excel），获取要素及其上下文信息（可选过滤材料名称）
        factors = parse_factors_file(factors_file, material_name)
        factor_names = [f['name'] for f in factors]
        print(f"\n✓ 成功解析文件，共 {len(factors)} 个要素: {', '.join(factor_names)}")
        
        # 2. 检查图片
        images = get_images_in_dir(current_dir)
        if not images:
            print("\n[错误] 当前目录下未发现图片附件，无法进行智能分析。")
            print("请添加至少一张图片文件（jpg/png/bmp）用于分析。")
            return
        
        print(f"\n✓ 发现 {len(images)} 张图片，将逐张分析所有图片")
        
        # 3. 获取 Qwen 客户端
        client = get_qwen_client()
        if not client:
            return
        
        # 4. 加载案例库并尝试匹配
        case_lib = load_case_library()
        print(f"\n✓ 案例库已加载，共 {len(case_lib.get('cases', []))} 个案例")
        
        matched_cases = match_cases_with_qwen(client, factors, case_lib)
        
        # 5. 根据匹配结果决定是使用案例库还是AI生成
        factors_with_rules = []
        need_ai_generation = []
        
        for idx, factor in enumerate(factors, 1):
            if factor['name'] in matched_cases:
                # 完全使用案例库中的规则（包括提取规则和格式要求）
                case = matched_cases[factor['name']]
                factors_with_rules.append({
                    'index': idx,
                    'name': factor['name'],
                    'extract_desc': factor.get('extract_desc', ''),
                    'rule_desc': factor.get('rule_desc', ''),
                    'rule': case['extraction_rule'],
                    'format': case['format_requirement'],
                    'case_matched': True  # 标记为案例库匹配
                })
                print(f"[案例库] ✓ '{factor['name']}' 完全使用案例库规则（提取规则+格式要求）")
            else:
                # 需要AI生成
                need_ai_generation.append(factor)
        
        # 6. 对未匹配的要素使用AI生成规则
        if need_ai_generation:
            print(f"\n[AI生成] {len(need_ai_generation)} 个要素需要AI生成规则")
            
            # 让 Qwen 分析所有图片并识别要素（合并结果）
            analysis_result = analyze_image_with_qwen(client, need_ai_generation, images)
            if not analysis_result:
                print("\n[错误] 图片分析失败，无法继续。")
                return
            
            # 让 Qwen 根据识别结果生成智能提取规则
            smart_rules = generate_smart_rules(client, need_ai_generation, analysis_result)
            if not smart_rules:
                print("\n[错误] 智能规则生成失败，无法继续。")
                return
            
            # 解析智能规则
            ai_generated_factors = parse_smart_rules(smart_rules)
            if not ai_generated_factors:
                print("\n[错误] 智能规则解析失败，无法继续。")
                return
            
            # 合并AI生成的规则
            factors_with_rules.extend(ai_generated_factors)
            
            print(f"\n✓ 成功生成 {len(ai_generated_factors)} 个AI智能提取规则")
        
        # 重新排序，保持原CSV顺序，并重新分配索引
        factor_order = {f['name']: i for i, f in enumerate(factors)}
        factors_with_rules.sort(key=lambda x: factor_order.get(x['name'], 999))
        
        # 重新分配索引
        for idx, factor in enumerate(factors_with_rules, 1):
            factor['index'] = idx
        
        # 统计格式使用情况
        case_format_count = sum(1 for f in factors_with_rules if f.get('case_matched', False))
        ai_format_count = len(factors_with_rules) - case_format_count
        
        print(f"\n✓ 共 {len(factors_with_rules)} 个提取规则准备就绪")
        print(f"  - 案例库规则: {len(matched_cases)}")
        print(f"  - AI生成规则: {len(need_ai_generation)}")
        print(f"  - 格式要求: 案例库 {case_format_count}, AI {ai_format_count}")
        
        # 7. 加载模板并组装提示词
        template_text = load_template(template_path)
        factors_text = generate_factors_text(factors_with_rules)
        final_prompt = build_prompt(template_text, factors_text)
        
        # 8. 保存提示词
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(final_prompt)
        print(f"\n✓ 智能提示词已保存至: {output_path}")
        
        # 9. 最终验证（对所有图片逐张验证）
        validation_success = validate_final_prompt(client, final_prompt, images)
        
        # 10. 如果验证成功且有AI生成的规则，保存到案例库
        if validation_success and need_ai_generation:
            # 只保存AI生成的规则到案例库
            ai_factors = [f for f in factors_with_rules if f['name'] in [nf['name'] for nf in need_ai_generation]]
            add_cases_to_library(case_lib, ai_factors, source="ai_generated")
        
        print("\n" + "="*60)
        print("✓ 智能提示词生成完成！")
        print("="*60)
            
    except Exception as e:
        print(f"\n[错误] 运行失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
