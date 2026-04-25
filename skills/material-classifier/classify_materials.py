import os
import sys
import json
import base64
import shutil
import re
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("[警告] 未安装 openpyxl，Excel 支持不可用。安装命令: pip install openpyxl")

try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("[警告] 未安装 PyMuPDF，PDF 支持不可用。安装命令: pip install pymupdf")


# ─────────────────────────────────────────────
# 路径辅助
# ─────────────────────────────────────────────

def get_skill_dir():
    """返回 skill 脚本所在目录"""
    return os.path.dirname(os.path.abspath(__file__))


# ─────────────────────────────────────────────
# 额外模型参数（来自桌面端设置）
# ─────────────────────────────────────────────

# DashScope 通过 OpenAI 兼容接口传递时，这些参数必须放在 extra_body 而非顶层 kwargs
_DASHSCOPE_BODY_PARAMS = {
    'enable_thinking', 'thinking_budget', 'translation_options',
    'vl_high_resolution_images', 'search_options',
}

def get_extra_params():
    """读取桌面端传入的额外模型参数，拆分为标准参数和 DashScope 专有参数(extra_body)"""
    raw = os.environ.get("CLASSIFY_EXTRA_PARAMS", "{}")
    try:
        all_params = json.loads(raw)
    except Exception:
        all_params = {}
    standard = {k: v for k, v in all_params.items() if k not in _DASHSCOPE_BODY_PARAMS}
    body = {k: v for k, v in all_params.items() if k in _DASHSCOPE_BODY_PARAMS}
    if body:
        standard['extra_body'] = body
    return standard


# ─────────────────────────────────────────────
# Qwen 客户端
# ─────────────────────────────────────────────

def get_qwen_client():
    """获取 Qwen API 客户端（与 doc-extract-prompt-gen 共用同一 key）"""
    try:
        from openai import OpenAI
    except ImportError:
        print("\n[错误] 缺少 openai 库，请执行: pip install openai")
        return None

    api_key = os.environ.get("DASHSCOPE_API_KEY")
    if not api_key:
        print("\n[错误] 未检测到 DASHSCOPE_API_KEY 环境变量")
        print("请设置环境变量: export DASHSCOPE_API_KEY='your-api-key'")
        return None

    from openai import OpenAI
    return OpenAI(
        api_key=api_key,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    )


# ─────────────────────────────────────────────
# CSV / Excel 解析
# ─────────────────────────────────────────────

def _is_excel_file(path):
    """通过魔数字节判断是否为 Excel (xlsx/xls) 文件，不依赖扩展名"""
    with open(path, 'rb') as f:
        magic = f.read(4)
    # xlsx/docx/zip: PK\x03\x04
    if magic[:4] == b'PK\x03\x04':
        return True
    # xls (BIFF): D0 CF 11 E0
    if magic[:4] == b'\xd0\xcf\x11\xe0':
        return True
    return False


def read_material_names(factors_path):
    """从 factors 文件（xlsx/xls/csv）读取材料名称列（B列，index=1）并去重，保留原始顺序。
    格式：A=事项名称, B=材料名称, ...
    通过魔数检测文件格式，兼容扩展名为 .csv 但实为 xlsx 的情况。
    读取所有行（包括无要素定义的行），清理首尾空格。"""
    if not os.path.exists(factors_path):
        raise FileNotFoundError(f"未找到 factors 文件: {factors_path}")

    use_excel = _is_excel_file(factors_path)

    if use_excel:
        if not EXCEL_SUPPORT:
            raise ImportError("读取 Excel 需要 openpyxl: pip install openpyxl")
        import tempfile, shutil as _shutil
        ext = os.path.splitext(factors_path)[1].lower()
        # openpyxl 拒绝 .csv 扩展名，若实为 xlsx 则临时复制为 .xlsx
        if ext not in ('.xlsx', '.xls', '.xlsm', '.xltx', '.xltm'):
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.close()
            _shutil.copy2(factors_path, tmp.name)
            load_path = tmp.name
            cleanup_tmp = True
        else:
            load_path = factors_path
            cleanup_tmp = False
        try:
            wb = load_workbook(load_path, read_only=True, data_only=True)
            ws = wb.active
            names = []
            seen = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 材料名称在 B 列（index=1），A 列是事项名称
                if not row or len(row) < 2 or not row[1]:
                    continue
                # 清理首尾空格并标准化
                val = str(row[1]).strip()
                # 过滤异常值：跳过多行文本、超长文本（>100字符）、包含明显说明性文字
                if not val or val in seen:
                    continue
                if '\n' in val or len(val) > 100:
                    continue
                if any(keyword in val for keyword in ['、', '①', '②', '③', '审查', '梳理', '说明', '参考', '例如']):
                    continue
                seen.add(val)
                names.append(val)
            wb.close()
        finally:
            if cleanup_tmp:
                os.unlink(tmp.name)
        
        # 打印读取到的材料类别用于调试
        print(f"[材料类别] 从 factors 文件读取到 {len(names)} 种材料")
        return names

    # 纯文本 CSV
    import csv
    names = []
    seen = set()
    with open(factors_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header
        for row in reader:
            # 材料名称在 B 列（index=1），A 列是事项名称
            if not row or len(row) < 2 or not row[1]:
                continue
            val = row[1].strip()
            # 过滤异常值：跳过多行文本、超长文本（>100字符）、包含明显说明性文字
            if not val or val in seen:
                continue
            if '\n' in val or len(val) > 100:
                continue
            if any(keyword in val for keyword in ['、', '①', '②', '③', '审查', '梳理', '说明', '参考', '例如']):
                continue
            seen.add(val)
            names.append(val)
    
    print(f"[材料类别] 从 factors 文件读取到 {len(names)} 种材料")
    return names


def find_factors_file(base_dir):
    """在指定目录查找 factors 文件，优先 xlsx"""
    for ext in ('.xlsx', '.xls', '.csv'):
        p = os.path.join(base_dir, f'factors{ext}')
        if os.path.exists(p):
            return p
    return None


# ─────────────────────────────────────────────
# 目录管理
# ─────────────────────────────────────────────

def ensure_classified_dirs(base_dir, material_names):
    """在 已分类材料/ 下为每种材料名称创建目录"""
    classified_root = os.path.join(base_dir, '已分类材料')
    os.makedirs(classified_root, exist_ok=True)
    created = []
    for name in material_names:
        folder = os.path.join(classified_root, name)
        if not os.path.exists(folder):
            os.makedirs(folder)
            created.append(name)
            print(f"[目录] 创建: 已分类材料/{name}/")
        else:
            print(f"[目录] 已存在: 已分类材料/{name}/")
    return classified_root, created


# ─────────────────────────────────────────────
# 模板读写
# ─────────────────────────────────────────────

def load_template(template_path, skill_default_name):
    """加载提示词模板，优先工作目录，备选 skill 目录"""
    if os.path.exists(template_path) and os.path.getsize(template_path) > 0:
        with open(template_path, 'r', encoding='utf-8') as f:
            return f.read()
    # 备选：skill 目录默认模板
    skill_path = os.path.join(get_skill_dir(), skill_default_name)
    if os.path.exists(skill_path):
        with open(skill_path, 'r', encoding='utf-8') as f:
            return f.read()
    raise FileNotFoundError(f"未找到模板: {template_path} 或 {skill_path}")


def save_template(path, content, use_timestamp=False):
    """保存模板文件，默认直接覆盖原文件，仅保留最新版本
    
    Args:
        path: 原始模板路径
        content: 模板内容
        use_timestamp: 是否生成带时间戳的新文件（默认False，直接覆盖）
    
    Returns:
        实际保存的文件路径
    """
    if use_timestamp:
        # 清理旧的优化版本文件
        dir_name = os.path.dirname(path)
        base_name = os.path.basename(path)
        name_without_ext = os.path.splitext(base_name)[0]
        ext = os.path.splitext(base_name)[1]
        
        # 删除所有旧的优化版本
        if os.path.isdir(dir_name):
            for f in os.listdir(dir_name):
                if f.startswith(f"{name_without_ext}_优化版_") and f.endswith(ext):
                    old_file = os.path.join(dir_name, f)
                    try:
                        os.remove(old_file)
                        print(f"[清理] 已删除旧版本: {f}")
                    except Exception as e:
                        print(f"[警告] 删除旧版本失败 {f}: {e}")
        
        # 生成带时间戳的新文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{name_without_ext}_优化版_{timestamp}{ext}"
        save_path = os.path.join(dir_name, new_name)
    else:
        # 直接覆盖原文件
        save_path = path
    
    with open(save_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"[模板] 已保存: {save_path}")
    return save_path


# ─────────────────────────────────────────────
# 图片工具
# ─────────────────────────────────────────────

def get_images_in_dir(directory):
    valid_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.pdf'}
    files = sorted([
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if Path(f).suffix.lower() in valid_exts
    ])
    return files


def encode_image(image_path):
    """对图片文件 base64 编码；PDF 则渲染首页为图片后编码"""
    ext = Path(image_path).suffix.lower()
    if ext == '.pdf':
        return encode_pdf_page(image_path)
    with open(image_path, 'rb') as f:
        return base64.b64encode(f.read()).decode('utf-8')


def encode_pdf_page(pdf_path, page_index=0, dpi=150):
    """将 PDF 指定页渲染为 PNG 并返回 base64 字符串"""
    if not PDF_SUPPORT:
        raise ImportError("需要 PyMuPDF 才能处理 PDF 文件，请执行: pip install pymupdf")
    doc = fitz.open(pdf_path)
    page = doc[min(page_index, len(doc) - 1)]
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
    doc.close()
    return base64.b64encode(pix.tobytes('png')).decode('utf-8')


def extract_json(text):
    """从 LLM 输出中提取 JSON 块"""
    if '```json' in text:
        start = text.find('```json') + 7
        end = text.find('```', start)
        return text[start:end].strip()
    if '```' in text:
        start = text.find('```') + 3
        end = text.find('```', start)
        return text[start:end].strip()
    # 尝试找到第一个 { 到最后一个 }
    start = text.find('{')
    end = text.rfind('}')
    if start != -1 and end != -1:
        return text[start:end + 1].strip()
    return text.strip()


def _extract_first_json_value(text):
    """尽力从文本中提取第一个可解析 JSON（对象或数组）"""
    if not text:
        return None
    decoder = json.JSONDecoder()
    for i, ch in enumerate(text):
        if ch not in '{[':
            continue
        try:
            value, _ = decoder.raw_decode(text[i:])
            return value
        except Exception:
            continue
    return None


def _normalize_step2_payload(data):
    """兼容不同返回格式，统一提取 classification_plan/summary"""
    if isinstance(data, list):
        return data, {}

    if not isinstance(data, dict):
        return None, None

    # 常见字段兼容：classification_plan / classificationPlan / plan
    plan = (
        data.get('classification_plan')
        or data.get('classificationPlan')
        or data.get('plan')
        or data.get('attachments_plan')
        or []
    )
    summary = data.get('summary') or {}

    # 某些模型会把 plan 作为字符串 JSON 返回
    if isinstance(plan, str):
        try:
            maybe = json.loads(plan)
            if isinstance(maybe, list):
                plan = maybe
        except Exception:
            pass

    if not isinstance(plan, list):
        return None, None
    if not isinstance(summary, dict):
        summary = {}
    return plan, summary


# ─────────────────────────────────────────────
# 步骤1：分类信息提取
# ─────────────────────────────────────────────

def run_step1_extraction(client, images, extract_template, material_names):
    """使用分类信息提取提示词对所有图片进行分类识别"""
    print(f"\n[步骤1] 分类信息提取，共 {len(images)} 张附件...")

    material_list_text = '\n'.join(f'- {name}' for name in material_names)
    prompt = extract_template.replace('$(material_list)', material_list_text)

    # 构建消息内容（文字 + 所有图片/PDF首页）
    content = [{"type": "text", "text": prompt}]
    for img_path in images:
        try:
            b64 = encode_image(img_path)
        except Exception as e:
            print(f"[警告] 无法编码文件 {os.path.basename(img_path)}: {e}，跳过")
            continue
        ext = Path(img_path).suffix.lower().lstrip('.')
        # PDF 渲染为 PNG
        mime = 'png' if ext == 'pdf' else ('jpeg' if ext in ('jpg', 'jpeg') else ext)
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/{mime};base64,{b64}"}
        })
        content.append({"type": "text", "text": f"[以上图片文件名: {os.path.basename(img_path)}]"})


    model_name = os.environ.get("CLASSIFY_MODEL_NAME", "qwen-vl-max")
    extra = get_extra_params()
    import time as _time
    _t0 = _time.time()
    try:
        response = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": content}],
            **extra
        )
        result = response.choices[0].message.content
        _elapsed = round(_time.time() - _t0, 2)
        print("[步骤1] 原始返回:")
        print(result[:800])
        print("__LLM_LOG__:" + json.dumps({
            "scene": "材料分类-步骤1", "model": model_name,
            "prompt_summary": prompt[:2000], "response_summary": result[:2000],
            "elapsed_s": _elapsed, "success": True
        }, ensure_ascii=False))
        return result
    except Exception as e:
        print(f"[错误] 步骤1 API 调用失败: {e}")
        print("__LLM_LOG__:" + json.dumps({
            "scene": "材料分类-步骤1", "model": model_name,
            "prompt_summary": prompt[:2000], "response_summary": "",
            "elapsed_s": None, "success": False, "error": str(e)
        }, ensure_ascii=False))
        return None


def parse_step1_result(raw_text):
    """解析步骤1 JSON 返回"""
    try:
        json_str = extract_json(raw_text)
        data = json.loads(json_str)
        attachments = data.get('attachments', [])
        print(f"[步骤1] 成功解析，共识别 {len(attachments)} 条分类记录")
        return attachments
    except Exception as e:
        print(f"[警告] 步骤1 结果解析失败: {e}")
        return None


# ─────────────────────────────────────────────
# 步骤2：附件归集
# ─────────────────────────────────────────────

def run_step2_aggregation(client, images, aggregate_template, material_names, step1_result_raw):
    """使用分类附件归集提示词生成最终归集方案"""
    print(f"\n[步骤2] 附件归集...")

    material_list_text = '\n'.join(f'- {name}' for name in material_names)
    prompt = (aggregate_template
              .replace('$(material_list)', material_list_text)
              .replace('$(classification_result)', step1_result_raw or '（步骤1结果为空）'))

    content = [{"type": "text", "text": prompt}]

    model_name = os.environ.get("CLASSIFY_MODEL_NAME", "qwen-vl-max")
    extra = get_extra_params()
    import time as _time
    _t0 = _time.time()
    try:
        response = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": content}],
            **extra
        )
        result = response.choices[0].message.content
        _elapsed = round(_time.time() - _t0, 2)
        print("[步骤2] 原始返回:")
        print(result[:800])
        print("__LLM_LOG__:" + json.dumps({
            "scene": "材料分类-步骤2", "model": model_name,
            "prompt_summary": prompt[:2000], "response_summary": result[:2000],
            "elapsed_s": _elapsed, "success": True
        }, ensure_ascii=False))
        return result
    except Exception as e:
        print(f"[错误] 步骤2 API 调用失败: {e}")
        print("__LLM_LOG__:" + json.dumps({
            "scene": "材料分类-步骤2", "model": model_name,
            "prompt_summary": prompt[:2000], "response_summary": "",
            "elapsed_s": None, "success": False, "error": str(e)
        }, ensure_ascii=False))
        return None


def parse_step2_result(raw_text):
    """解析步骤2 JSON 返回"""
    try:
        json_str = extract_json(raw_text)
        data = None
        try:
            data = json.loads(json_str)
        except Exception:
            # 模型输出夹带说明文字时，尝试从整段文本抽取第一个 JSON 值
            data = _extract_first_json_value(raw_text)

        plan, summary = _normalize_step2_payload(data)
        if plan is None:
            raise ValueError("missing classification_plan")

        print(f"[步骤2] 成功解析，共 {len(plan)} 条归集方案")
        return plan, summary
    except Exception as e:
        print(f"[警告] 步骤2 结果解析失败: {e}")
        if raw_text:
            text = str(raw_text)
            # 记录截断内容，便于排查模型格式漂移
            print(f"[步骤2] 原始输出片段: {text[:400]}")
        return None, None


# ─────────────────────────────────────────────
# AI 优化提示词
# ─────────────────────────────────────────────

def optimize_prompt_with_qwen(client, prompt_type, current_template, material_names,
                               step1_result, step2_result, iteration, issues):
    """让 Qwen 根据当前运行结果优化提示词模板，使用设置中的造物主提示词"""
    print(f"\n[优化] 正在优化 {prompt_type}（第 {iteration} 轮）...")

    material_list_text = '\n'.join(f'- {name}' for name in material_names)

    # 从环境变量读取造物主提示词（由桌面端设置传入）
    god_prompt = os.environ.get("CLASSIFY_GOD_PROMPT", "").strip()

    if prompt_type == '分类信息提取提示词':
        task_context = f"""请根据以下信息优化"分类信息提取提示词模板"。

## 当前模板
```
{current_template}
```

## 可用材料类别
{material_list_text}

## 当前模板运行结果
```json
{json.dumps(step1_result, ensure_ascii=False, indent=2) if step1_result else '未能成功解析'}
```

## 存在的问题
{issues if issues else '暂无明确问题，请根据结果质量进行优化'}

## 优化要求
1. 模板必须保留 `$(material_list)` 占位符，用于插入材料类别列表
2. 输出 JSON 格式必须保持不变（attachments 数组，每项含 file_name/material_type/key_info/reason）
3. 优化分类识别的准确性，使 LLM 能更精确地识别每个附件属于哪种材料
4. 针对存在的问题给出针对性改进
5. 只返回优化后的完整模板文本，不要添加额外说明

请直接返回优化后的模板内容（纯文本，不要包含 markdown 代码块标记）："""

    else:  # 分类附件归集提示词
        task_context = f"""请根据以下信息优化"分类附件归集提示词模板"。

## 当前模板
```
{current_template}
```

## 可用材料类别
{material_list_text}

## 步骤1分类提取结果（输入）
```json
{json.dumps(step1_result, ensure_ascii=False, indent=2) if step1_result else '未能成功解析'}
```

## 步骤2归集方案结果（当前模板输出）
```json
{json.dumps(step2_result, ensure_ascii=False, indent=2) if step2_result else '未能成功解析'}
```

## 存在的问题
{issues if issues else '暂无明确问题，请根据结果质量进行优化'}

## 优化要求
1. 模板必须保留 `$(material_list)` 占位符（材料目录列表）和 `$(classification_result)` 占位符（步骤1结果）
2. 输出 JSON 格式必须保持不变（classification_plan 数组 + summary 对象）
3. 优化归集方案的准确性和完整性
4. 针对存在的问题给出针对性改进
5. 只返回优化后的完整模板文本，不要添加额外说明

请直接返回优化后的模板内容（纯文本，不要包含 markdown 代码块标记）："""

    # 构建消息：若有造物主提示词则作为 system 消息
    if god_prompt:
        messages = [
            {"role": "system", "content": god_prompt},
            {"role": "user", "content": task_context}
        ]
        print(f"[优化] 使用造物主提示词（{len(god_prompt)} 字符）")
        print(f"[优化] 造物主提示词内容:\n{god_prompt}\n---")
    else:
        messages = [{"role": "user", "content": "你是一个专业的提示词优化专家。" + task_context}]
        print("[优化] 未使用造物主提示词（为空）")

    model_name = os.environ.get("CLASSIFY_MODEL_NAME", "qwen-plus")
    extra = get_extra_params()
    import time as _time
    _t0 = _time.time()
    try:
        response = client.chat.completions.create(
            model=model_name,
            messages=messages,
            **extra
        )
        optimized = response.choices[0].message.content.strip()
        _elapsed = round(_time.time() - _t0, 2)
        # 去除可能存在的 markdown 代码块包裹
        if optimized.startswith('```') and optimized.endswith('```'):
            lines = optimized.split('\n')
            optimized = '\n'.join(lines[1:-1])
        print(f"[优化] 完成，新模板长度: {len(optimized)} 字符")
        print("__LLM_LOG__:" + json.dumps({
            "scene": f"材料分类-优化({prompt_type})", "model": model_name,
            "prompt_summary": task_context[:2000], "response_summary": optimized[:2000],
            "elapsed_s": _elapsed, "success": True
        }, ensure_ascii=False))
        return optimized
    except Exception as e:
        print(f"[错误] 提示词优化失败: {e}")
        print("__LLM_LOG__:" + json.dumps({
            "scene": f"材料分类-优化({prompt_type})", "model": model_name,
            "prompt_summary": task_context[:2000], "response_summary": "",
            "elapsed_s": None, "success": False, "error": str(e)
        }, ensure_ascii=False))
        return current_template


# ─────────────────────────────────────────────
# 质量评估
# ─────────────────────────────────────────────

def evaluate_step1(attachments, image_files, material_names):
    """评估步骤1结果质量，返回 (是否通过, 问题列表)"""
    issues = []
    if not attachments:
        return False, ["步骤1未返回任何分类记录"]

    image_basenames = {os.path.basename(p) for p in image_files}
    recognized_files = {a.get('file_name', '') for a in attachments}

    # 检查是否所有图片都有分类结果
    missing = image_basenames - recognized_files
    if missing:
        issues.append(f"以下附件未被识别: {', '.join(sorted(missing))}")

    # 检查 material_type 是否在有效列表内
    invalid_types = [
        a.get('material_type', '') for a in attachments
        if a.get('material_type', '') not in material_names and a.get('material_type', '') != '其他'
    ]
    if invalid_types:
        issues.append(f"存在无效材料类型: {set(invalid_types)}")

    # 检查必要字段
    for a in attachments:
        for field in ('file_name', 'material_type', 'reason'):
            if not a.get(field):
                issues.append(f"附件 {a.get('file_name', '?')} 缺少字段: {field}")

    passed = len(issues) == 0
    return passed, issues


def evaluate_step2(plan, summary, image_files):
    """评估步骤2结果质量，返回 (是否通过, 问题列表)"""
    issues = []
    if not plan:
        return False, ["步骤2未返回任何归集方案"]

    image_basenames = {os.path.basename(p) for p in image_files}
    planned_files = {p.get('file_name', '') for p in plan}

    missing = image_basenames - planned_files
    if missing:
        issues.append(f"以下附件未在归集方案中: {', '.join(sorted(missing))}")

    for p in plan:
        if not p.get('target_folder'):
            issues.append(f"附件 {p.get('file_name', '?')} 缺少目标目录")
        if p.get('confidence') not in ('high', 'medium', 'low'):
            issues.append(f"附件 {p.get('file_name', '?')} 置信度格式有误")

    passed = len(issues) == 0
    return passed, issues


# ─────────────────────────────────────────────
# 执行归集（复制文件到目标目录）
# ─────────────────────────────────────────────

def execute_classification(plan, unclassified_dir, classified_root, material_names):
    """根据归集方案将附件复制到对应目录"""
    print(f"\n[归集] 开始执行文件归集...")
    success_count = 0
    fail_list = []

    # 确保所有材料目录已存在
    for name in material_names:
        os.makedirs(os.path.join(classified_root, name), exist_ok=True)
    other_dir = os.path.join(classified_root, '其他')
    os.makedirs(other_dir, exist_ok=True)

    for item in plan:
        file_name = item.get('file_name', '')
        target_folder = item.get('target_folder', '其他')
        confidence = item.get('confidence', 'low')

        src = os.path.join(unclassified_dir, file_name)
        if not os.path.exists(src):
            print(f"[归集] ✗ 源文件不存在: {file_name}")
            fail_list.append(file_name)
            continue

        # 目标目录：如果不在材料列表中，归入"其他"
        if target_folder not in material_names:
            target_folder = '其他'

        dst_dir = os.path.join(classified_root, target_folder)
        os.makedirs(dst_dir, exist_ok=True)
        dst = os.path.join(dst_dir, file_name)

        try:
            shutil.copy2(src, dst)
            print(f"[归集] ✓ {file_name} → 已分类材料/{target_folder}/ (置信度: {confidence})")
            success_count += 1
        except Exception as e:
            print(f"[归集] ✗ 复制失败 {file_name}: {e}")
            fail_list.append(file_name)

    print(f"\n[归集] 完成：成功 {success_count} 个，失败 {len(fail_list)} 个")
    if fail_list:
        print(f"[归集] 失败文件: {fail_list}")
    return success_count, fail_list


# ─────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────

def main():
    # 解析命令行参数
    if len(sys.argv) > 1:
        base_dir = os.path.abspath(sys.argv[1])
    else:
        # 默认使用 分类材料集 目录（相对 Auto-Prompt 根目录）
        script_dir = get_skill_dir()
        base_dir = os.path.normpath(
            os.path.join(script_dir, '..', '..', '..', '分类材料集')
        )

    max_iterations = int(sys.argv[2]) if len(sys.argv) > 2 else 3

    print("=" * 60)
    print("材料分类提示词自动优化工具")
    print("=" * 60)
    print(f"[配置] 工作目录: {base_dir}")
    print(f"[配置] 最大优化轮次: {max_iterations}")

    # ── 1. 校验目录结构
    unclassified_dir = None
    for name in ['待分类材料', '待分类']:
        candidate = os.path.join(base_dir, name)
        if os.path.isdir(candidate):
            unclassified_dir = candidate
            break
    if not unclassified_dir:
        print(f"[错误] 待分类材料目录不存在: {base_dir}/待分类材料 或 {base_dir}/待分类")
        return 1

    # ── 2. 读取材料名称
    factors_path = find_factors_file(base_dir)
    if not factors_path:
        print(f"[错误] 未找到 factors.csv/xlsx 文件: {base_dir}")
        return 1
    material_names = read_material_names(factors_path)
    print(f"\n✓ 材料类别（{len(material_names)} 种）: {', '.join(material_names)}")

    # ── 3. 建立已分类材料目录
    classified_root, created = ensure_classified_dirs(base_dir, material_names)

    # ── 4. 获取待分类附件列表
    image_files = get_images_in_dir(unclassified_dir)
    if not image_files:
        print(f"[错误] 待分类材料目录中未找到图片附件: {unclassified_dir}")
        return
    print(f"\n✓ 待分类附件：{len(image_files)} 张")
    for img in image_files:
        print(f"  - {os.path.basename(img)}")

    # ── 5. 加载提示词模板
    extract_tmpl_path = os.path.join(base_dir, '分类信息提取提示词模板.txt')
    aggregate_tmpl_path = os.path.join(base_dir, '分类附件归集提示词模板.txt')

    extract_template = load_template(extract_tmpl_path, '分类信息提取提示词模板.txt')
    aggregate_template = load_template(aggregate_tmpl_path, '分类附件归集提示词模板.txt')
    print(f"\n✓ 提示词模板已加载")

    # ── 6. 获取 Qwen 客户端
    client = get_qwen_client()
    if not client:
        return 1

    # ── 7. 迭代优化循环
    step1_raw = None
    step1_attachments = None
    step2_raw = None
    step2_plan = None
    step2_summary = None
    
    # 记录最终保存的文件路径
    final_extract_path = extract_tmpl_path
    final_aggregate_path = aggregate_tmpl_path
    iterations_run = 0
    extract_was_optimized = False
    aggregate_was_optimized = False

    for iteration in range(1, max_iterations + 1):
        iterations_run = iteration
        print(f"\n{'─'*60}")
        print(f"[轮次 {iteration}/{max_iterations}]")
        print(f"{'─'*60}")

        # 步骤1：分类信息提取
        step1_raw = run_step1_extraction(client, image_files, extract_template, material_names)
        if not step1_raw:
            print("[错误] 步骤1执行失败，终止")
            return 1

        step1_attachments = parse_step1_result(step1_raw)
        s1_pass, s1_issues = evaluate_step1(
            step1_attachments or [], image_files, material_names
        )
        print(f"[评估] 步骤1 {'✓ 通过' if s1_pass else '✗ 未通过'}")
        if s1_issues:
            for issue in s1_issues:
                print(f"  - {issue}")

        # 步骤2：附件归集
        step2_raw = run_step2_aggregation(
            client, image_files, aggregate_template,
            material_names, step1_raw
        )
        if not step2_raw:
            print("[错误] 步骤2执行失败，终止")
            return 1

        step2_plan, step2_summary = parse_step2_result(step2_raw)
        s2_pass, s2_issues = evaluate_step2(
            step2_plan or [], step2_summary, image_files
        )
        print(f"[评估] 步骤2 {'✓ 通过' if s2_pass else '✗ 未通过'}")
        if s2_issues:
            for issue in s2_issues:
                print(f"  - {issue}")

        # 若两步均通过，退出循环
        if s1_pass and s2_pass:
            print(f"\n✓ 第 {iteration} 轮两步均通过，无需继续优化")
            break

        # 还有剩余轮次才优化
        if iteration < max_iterations:
            if not s1_pass:
                extract_template = optimize_prompt_with_qwen(
                    client, '分类信息提取提示词',
                    extract_template, material_names,
                    step1_attachments, None,
                    iteration, '\n'.join(s1_issues)
                )
                final_extract_path = save_template(extract_tmpl_path, extract_template, use_timestamp=False)
                extract_was_optimized = True

            if not s2_pass:
                aggregate_template = optimize_prompt_with_qwen(
                    client, '分类附件归集提示词',
                    aggregate_template, material_names,
                    step1_attachments, step2_plan,
                    iteration, '\n'.join(s2_issues)
                )
                final_aggregate_path = save_template(aggregate_tmpl_path, aggregate_template, use_timestamp=False)
                aggregate_was_optimized = True
        else:
            print(f"\n[优化] 已达最大轮次 {max_iterations}，保存当前最优提示词")
            # 最后一轮也保存（即便未完全通过）
            final_extract_path = save_template(extract_tmpl_path, extract_template, use_timestamp=False)
            final_aggregate_path = save_template(aggregate_tmpl_path, aggregate_template, use_timestamp=False)
            extract_was_optimized = True
            aggregate_was_optimized = True

    # 无论是否触发优化，都将本次实际使用的提示词落盘到工作区，
    # 便于下载结果包和后续追溯。
    final_extract_path = save_template(extract_tmpl_path, extract_template, use_timestamp=False)
    final_aggregate_path = save_template(aggregate_tmpl_path, aggregate_template, use_timestamp=False)

    # ── 8. 输出最终优化后的提示词
    print(f"\n{'='*60}")
    print("最终优化提示词")
    print(f"{'='*60}")

    print("\n【分类信息提取提示词】")
    print("─" * 40)
    print(extract_template)

    print("\n【分类附件归集提示词】")
    print("─" * 40)
    print(aggregate_template)

    # ── 9. 自动执行归集（若步骤2有结果）
    if step2_plan:
        print("\n[归集] 开始执行文件归集...")
        execute_classification(step2_plan, unclassified_dir, classified_root, material_names)
    else:
        print("\n[警告] 步骤2未生成有效归集方案，跳过文件归集")
        return 1

    # ── 10. 保存运行报告
    def _prompt_source(was_optimized, iters):
        if not was_optimized:
            return "原始模板"
        return f"AI优化（共{iters}轮）"

    report = {
        "run_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "base_dir": base_dir,
        "material_names": material_names,
        "image_count": len(image_files),
        "iterations_run": iterations_run,
        "extract_prompt_source": _prompt_source(extract_was_optimized, iterations_run),
        "aggregate_prompt_source": _prompt_source(aggregate_was_optimized, iterations_run),
        "step1_result": step1_attachments,
        "step2_plan": step2_plan,
        "step2_summary": step2_summary,
    }
    report_path = os.path.join(base_dir, 'classification_report.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n✓ 运行报告已保存: {report_path}")

    print("\n" + "=" * 60)
    print("✓ 材料分类完成！")
    print(f"  - 分类信息提取提示词: {final_extract_path}")
    print(f"  - 分类附件归集提示词: {final_aggregate_path}")
    print(f"  - 已分类材料目录:     {classified_root}")
    print(f"  - 运行报告:           {report_path}")
    print("=" * 60)
    return 0


def test_prompt(work_dir, prompt_type, prompt_content):
    """单提示词调优测试：用给定提示词内容跑一次推理，返回 JSON 结果到 stdout
    
    Args:
        work_dir: 工作目录（需包含待分类材料/）
        prompt_type: 'extract' 或 'aggregate'
        prompt_content: 提示词文本内容
    """
    base_dir = os.path.abspath(work_dir)
    
    # 查找待分类材料
    unclassified_dir = None
    for name in ['待分类材料', '待分类']:
        p = os.path.join(base_dir, name)
        if os.path.exists(p):
            unclassified_dir = p
            break
    if not unclassified_dir:
        print(json.dumps({"error": "未找到待分类材料目录"}, ensure_ascii=False))
        return

    image_files = get_images_in_dir(unclassified_dir)
    if not image_files:
        print(json.dumps({"error": "待分类材料目录中无可处理文件"}, ensure_ascii=False))
        return

    # 读取材料名称
    factors_path = find_factors_file(base_dir)
    if not factors_path:
        print(json.dumps({"error": "未找到 factors 文件"}, ensure_ascii=False))
        return
    material_names = read_material_names(factors_path)

    client = get_qwen_client()
    if not client:
        print(json.dumps({"error": "无法初始化 Qwen 客户端，请检查 API Key"}, ensure_ascii=False))
        return

    print(f"[测试] 开始测试 {'分类信息提取' if prompt_type == 'extract' else '附件归集'} 提示词...")
    print(f"[测试] 待处理文件: {len(image_files)} 个")

    if prompt_type == 'extract':
        step1_raw = run_step1_extraction(client, image_files, prompt_content, material_names)
        if not step1_raw:
            print(json.dumps({"error": "步骤1执行失败"}, ensure_ascii=False))
            return
        attachments = parse_step1_result(step1_raw)
        s1_pass, s1_issues = evaluate_step1(attachments or [], image_files, material_names)
        result = {
            "type": "extract",
            "pass": s1_pass,
            "issues": s1_issues,
            "attachments": attachments or [],
            "raw": step1_raw,
        }
        print("TEST_RESULT_JSON:" + json.dumps(result, ensure_ascii=False))
    else:
        # aggregate 需要先跑 step1
        aggregate_tmpl_path = os.path.join(base_dir, '分类信息提取提示词模板.txt')
        extract_template = load_template(aggregate_tmpl_path, '默认分类信息提取模板.txt')
        step1_raw = run_step1_extraction(client, image_files, extract_template, material_names)
        if not step1_raw:
            print(json.dumps({"error": "step1 预处理失败"}, ensure_ascii=False))
            return
        step2_raw = run_step2_aggregation(client, image_files, prompt_content, material_names, step1_raw)
        if not step2_raw:
            print(json.dumps({"error": "步骤2执行失败"}, ensure_ascii=False))
            return
        step2_plan, step2_summary = parse_step2_result(step2_raw)
        s2_pass, s2_issues = evaluate_step2(step2_plan or [], step2_summary, image_files)
        result = {
            "type": "aggregate",
            "pass": s2_pass,
            "issues": s2_issues,
            "plan": step2_plan or [],
            "summary": step2_summary,
            "raw": step2_raw,
        }
        print("TEST_RESULT_JSON:" + json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument('--test-prompt', choices=['extract', 'aggregate'], default=None)
    parser.add_argument('--prompt-file', default=None)
    # positional: work_dir [max_rounds]
    parser.add_argument('args', nargs='*')
    parsed, _ = parser.parse_known_args()

    if parsed.test_prompt:
        if not parsed.args or not parsed.prompt_file:
            print(json.dumps({"error": "--test-prompt 需要 work_dir 和 --prompt-file"}, ensure_ascii=False))
            raise SystemExit(2)
        else:
            work_dir = parsed.args[0]
            with open(parsed.prompt_file, 'r', encoding='utf-8') as f:
                prompt_content = f.read()
            test_prompt(work_dir, parsed.test_prompt, prompt_content)
            raise SystemExit(0)
    else:
        raise SystemExit(main())
