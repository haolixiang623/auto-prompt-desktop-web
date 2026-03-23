#!/usr/bin/env python3
"""
要素信息录入JSON生成器 v2
输出符合导入规范的 {carriername, factors, promptGroups} 格式。

分组策略（按优先级）：
  1. 材料目录中有多个提示词 TXT 文件 → 每个 TXT 文件对应一个 promptGroup，
     组内要素 = 该 TXT 文件中出现的要素。
  2. 只有一个 TXT 文件 → 按 group_size（默认4）将所有要素平分为多组。
  3. 无 TXT 文件 → 仅生成 factors，不生成 promptGroups（系统自动归入默认组合）。

用法:
  python3 generate_factor_json.py <材料集目录> [--group-size N]
"""

import os
import sys
import json
import re
import math
import openpyxl
from pathlib import Path


# ─────────────────────────── Excel 读取 ────────────────────────────

def read_factors_from_excel(excel_path):
    """从 factors.xlsx 读取要素信息，自动适配两种列格式：

    格式A（简单格式）：A=材料名称, B=要素名称, C=要素用途, D=要素类型
    格式B（扩展格式）：A=事项名称, B=材料名称, D=要素字段名称, G=要素提取说明
      - 标题行中 A 列内容包含「事项」或 B 列包含「材料名称」时识别为格式B
      - 材料名称跨行合并，向下继承非空值

    Returns:
        dict: {材料名称: [(要素名称, 要素用途, 要素类型), ...]}
              要素类型默认 "1"（文本）
    """
    try:
        with open(excel_path, 'rb') as f:
            magic = f.read(4)
        if magic[:2] != b'PK':
            print(f"[错误] 不支持的文件格式: {excel_path}")
            return {}

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        factors_dict = {}

        # ── 读取标题行，自动判断格式 ──
        header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        # 格式B 特征：A1 含「事项」或 B1 = 「材料名称」
        is_extended = (
            ("事项" in header[0]) or
            (len(header) > 1 and "材料名称" in header[1])
        )

        if is_extended:
            # 格式B：B=材料名称(col1), D=要素字段名称(col3), G=要素提取说明(col6)
            # 用 col index (0-based): material=1, factor=3, usage=6
            IDX_MATERIAL = 1
            IDX_FACTOR   = 3
            IDX_USAGE    = 6
            IDX_TYPE     = -1  # 无要素类型列，默认 "1"
            print(f"[信息] 检测到扩展格式 Excel（B=材料名称, D=要素字段名称）")
        else:
            # 格式A：A=材料名称(col0), B=要素名称(col1), C=要素用途(col2), D=要素类型(col3)
            IDX_MATERIAL = 0
            IDX_FACTOR   = 1
            IDX_USAGE    = 2
            IDX_TYPE     = 3
            print(f"[信息] 检测到简单格式 Excel（A=材料名称, B=要素名称）")

        current_material = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 向下继承材料名称（合并单元格）
            raw_material = row[IDX_MATERIAL] if IDX_MATERIAL < len(row) else None
            if raw_material:
                s = str(raw_material).strip()
                # 过滤说明性内容（多行文字、含换行的注释块）
                if s and '\n' not in s and len(s) < 60:
                    current_material = s

            if not current_material:
                continue

            factor_name = str(row[IDX_FACTOR]).strip() if IDX_FACTOR < len(row) and row[IDX_FACTOR] else ""
            factor_usage = str(row[IDX_USAGE]).strip() if IDX_USAGE < len(row) and row[IDX_USAGE] else ""
            factor_type = "1"
            if IDX_TYPE >= 0 and IDX_TYPE < len(row) and row[IDX_TYPE]:
                factor_type = str(row[IDX_TYPE]).strip() or "1"

            if not factor_name:
                continue
            # 过滤要素名称中明显的说明性内容
            if '\n' in factor_name or len(factor_name) > 50:
                continue

            factors_dict.setdefault(current_material, [])
            factors_dict[current_material].append((factor_name, factor_usage, factor_type))

        return factors_dict

    except Exception as e:
        print(f"[错误] 读取Excel文件失败: {e}")
        return {}


# ─────────────────────────── TXT 提取 ────────────────────────────

def extract_prompt_rules(txt_path):
    """从提示词 TXT 文件中提取 {要素名称: 识别提示词} 字典，
    同时返回文件中要素的出现顺序列表（用于分组时确定组成员）。

    Returns:
        (dict, list): ({要素名称: 识别提示词}, [要素名称有序列表])
    """
    if not os.path.exists(txt_path):
        return {}, []

    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 兼容 "## 1.要素名称" 和 "## 1、要素名称" 两种格式
        pattern = r'##\s*\d+[\.、]\s*(.+?)\n(.*?)(?=\n##\s*\d+[\.、]|\n#\s|\Z)'
        matches = re.findall(pattern, content, re.DOTALL)

        prompt_dict = {}
        ordered_names = []
        for factor_name, rule_text in matches:
            factor_name = factor_name.strip()
            rule_text = rule_text.strip()
            # 移除格式说明后缀
            rule_text = re.sub(r'[，,]\s*(保持原格式|返回数组格式.*)$', '', rule_text, flags=re.MULTILINE)
            rule_text = rule_text.strip()
            if factor_name:
                prompt_dict[factor_name] = rule_text
                ordered_names.append(factor_name)

        return prompt_dict, ordered_names

    except Exception as e:
        print(f"[错误] 读取提示词文件失败 {txt_path}: {e}")
        return {}, []


def extract_group_prompt_template(txt_path):
    """从提示词 TXT 文件中提取整体 prompt_template（识别指令部分）。
    取"识别要素列表及规则"之前的主提示词区域，用 $(factors) 替换要素列表占位。
    如果找不到则返回空字符串。
    """
    if not os.path.exists(txt_path):
        return ""
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            content = f.read()
        # 取第一个 # 识别要素列表 之前的内容作为 template
        m = re.search(r'^(.*?)(?=\n#\s*识别要素列表|\Z)', content, re.DOTALL)
        if m:
            tmpl = m.group(1).strip()
            if tmpl:
                return tmpl + "\n$(factors)"
    except Exception:
        pass
    return ""


# ─────────────────────────── 分组构建 ────────────────────────────

def build_prompt_groups_from_txts(material_dir, factor_names, txt_files):
    """多 TXT 文件场景：每个 TXT 对应一个 promptGroup。

    组内要素 = TXT 中出现的 factor_names 交集（保持 Excel 顺序）。
    未被任何 TXT 覆盖的要素不单独建组（由系统归入默认组）。

    Returns:
        (merged_prompt_dict, groups)
        merged_prompt_dict: {要素名称: 识别提示词}（全量合并）
        groups: [{"groupname": str, "grouptype": "2", "ordernum": int,
                  "prompt_template": str, "factors": [str]}]
    """
    merged_prompt_dict = {}
    groups = []
    covered_factors = set()

    for idx, txt_file in enumerate(sorted(txt_files), 1):
        txt_path = os.path.join(material_dir, txt_file)
        prompt_dict, _ = extract_prompt_rules(txt_path)
        merged_prompt_dict.update(prompt_dict)

        # 组内要素 = Excel 要素 ∩ 此 TXT 中出现的要素（保持 Excel 顺序）
        group_factors = [f for f in factor_names if f in prompt_dict]
        covered_factors.update(group_factors)

        # 用文件名推导组名（去掉后缀和通用词）
        group_name = re.sub(r'[-_]?提示词.*$', '', txt_file.replace('.txt', '').strip())
        group_name = group_name or f"分组{idx}"

        template = extract_group_prompt_template(txt_path)

        groups.append({
            "groupname": group_name,
            "grouptype": "2",
            "ordernum": idx,
            "prompt_template": template,
            "modelguid": "",
            "factors": group_factors,
        })

    return merged_prompt_dict, groups


def build_prompt_groups_by_size(factor_names, prompt_dict, group_size, txt_path):
    """单 TXT 文件场景：按 group_size 将要素平分为多组。

    Returns:
        groups: [{"groupname": str, "grouptype": "2", ...}]
    """
    total = len(factor_names)
    num_groups = max(1, math.ceil(total / group_size))
    template = extract_group_prompt_template(txt_path) if txt_path else ""
    groups = []

    for i in range(num_groups):
        chunk = factor_names[i * group_size: (i + 1) * group_size]
        groups.append({
            "groupname": f"分组{i + 1}",
            "grouptype": "2",
            "ordernum": i + 1,
            "prompt_template": template,
            "modelguid": "",
            "factors": chunk,
        })

    return groups


# ─────────────────────────── 主体生成 ────────────────────────────

def generate_import_json(material_name, factors_info, material_dir, group_size=4):
    """为单个材料生成符合导入规范的 JSON 对象。

    Args:
        material_name: 材料名称（= carriername）
        factors_info: [(要素名称, 要素用途, 要素类型), ...]
        material_dir: 材料目录路径
        group_size: 单 TXT 时每组最多要素数

    Returns:
        dict: {carriername, factors, promptGroups}
    """
    factor_names = [f[0] for f in factors_info]

    # ── 扫描提示词 TXT 文件 ──
    txt_files = sorted([
        f for f in os.listdir(material_dir)
        if f.endswith('.txt') and '提示词' in f
    ])

    prompt_dict = {}
    groups = []

    if len(txt_files) > 1:
        print(f"  [信息] 发现 {len(txt_files)} 个提示词文件 → 每文件对应一个分组")
        prompt_dict, groups = build_prompt_groups_from_txts(material_dir, factor_names, txt_files)

    elif len(txt_files) == 1:
        txt_path = os.path.join(material_dir, txt_files[0])
        print(f"  [信息] 读取提示词: {txt_files[0]}")
        prompt_dict, _ = extract_prompt_rules(txt_path)
        print(f"  [信息] 提取到 {len(prompt_dict)} 个要素识别规则")
        print(f"  [信息] 按每组 {group_size} 个要素自动分组")
        groups = build_prompt_groups_by_size(factor_names, prompt_dict, group_size, txt_path)

    else:
        print(f"  [警告] 未找到提示词文件，按每组 {group_size} 个要素自动分组（prompt_template 留空）")
        groups = build_prompt_groups_by_size(factor_names, {}, group_size, None)

    # ── 构建 factors 数组 ──
    factors_list = []
    for idx, (factor_name, factor_usage, factor_type) in enumerate(factors_info, 1):
        factors_list.append({
            "factorname":    factor_name,
            "factortype":    factor_type or "1",
            "factor_prompt": prompt_dict.get(factor_name, ""),
            "factoruse":     factor_usage,
            "ordernum":      idx,
            "remark":        "",
            "is_usermsg":    "0",
            "isrecollect":   "0",
            "factor_trans":  "",
        })

    result = {
        "carriername":  material_name,
        "factors":      factors_list,
        "promptGroups": groups,
    }

    return result


# ─────────────────────────── 入口 ────────────────────────────────

def main(work_dir, group_size=4):
    work_dir = os.path.abspath(work_dir)

    if not os.path.isdir(work_dir):
        print(f"[错误] 工作目录不存在: {work_dir}")
        sys.exit(1)

    # 查找 factors.xlsx
    factors_path = None
    for fname in ['factors.xlsx', 'factors.csv']:
        p = os.path.join(work_dir, fname)
        if os.path.exists(p):
            factors_path = p
            break

    if not factors_path:
        print("[错误] 未找到 factors.xlsx 或 factors.csv 文件")
        sys.exit(1)

    print(f"[信息] 读取要素定义: {factors_path}")
    factors_dict = read_factors_from_excel(factors_path)

    if not factors_dict:
        print("[错误] 未读取到任何要素信息")
        sys.exit(1)

    print(f"[信息] 共读取 {len(factors_dict)} 个材料类别\n")

    results = []
    for material_name, factors_info in factors_dict.items():
        print(f"[处理] 材料: {material_name}")

        material_dir = os.path.join(work_dir, material_name)
        if not os.path.isdir(material_dir):
            print(f"  [警告] 材料目录不存在，跳过: {material_dir}")
            results.append({"material": material_name, "success": False,
                            "error": "材料目录不存在"})
            continue

        try:
            import_json = generate_import_json(
                material_name, factors_info, material_dir, group_size
            )

            output_path = os.path.join(material_dir, f"{material_name}--要素信息录入.json")
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(import_json, f, ensure_ascii=False, indent=2)

            factor_count = len(import_json["factors"])
            group_count  = len(import_json["promptGroups"])
            print(f"  [完成] 已生成: {output_path}")
            print(f"         要素: {factor_count} 个 · 分组: {group_count} 个\n")

            results.append({
                "material": material_name,
                "success": True,
                "output": output_path,
                "factor_count": factor_count,
                "group_count": group_count,
            })

        except Exception as e:
            print(f"  [错误] 生成失败: {e}\n")
            results.append({"material": material_name, "success": False, "error": str(e)})

    ok = sum(1 for r in results if r["success"])
    print(f"[完成] 所有材料处理完毕：成功 {ok} / {len(results)} 个")

    # 输出结构化结果供桌面端解析
    print("RESULTS_JSON:" + json.dumps(results, ensure_ascii=False))


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='要素信息录入JSON生成器')
    parser.add_argument('work_dir', help='材料集目录路径')
    parser.add_argument('--group-size', type=int, default=4,
                        help='单TXT文件时每个分组的最大要素数（默认4）')
    args = parser.parse_args()
    main(args.work_dir, args.group_size)
