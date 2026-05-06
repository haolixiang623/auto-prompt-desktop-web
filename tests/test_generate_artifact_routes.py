from pathlib import Path
import io
from zipfile import ZipFile

import httpx
import openpyxl
import pytest

from pyserver.app import main
from pyserver.app.factor_prompt_artifacts import build_preview_prompt
from pyserver.app.main import app


async def login(client: httpx.AsyncClient) -> str:
    response = await client.post(
        "/api/auth/login",
        json={"username": "admin", "password": "admin123456"},
    )
    response.raise_for_status()
    return response.json()["token"]


def _bootstrap_env(tmp_path, monkeypatch):
    repo_root = tmp_path / "repo"
    data_dir = repo_root / ".runtime-data"
    repo_root.mkdir(parents=True, exist_ok=True)
    data_dir.mkdir(parents=True, exist_ok=True)
    monkeypatch.setenv("AUTO_PROMPT_REPO_ROOT", str(repo_root))
    monkeypatch.setenv("AUTO_PROMPT_DATA_DIR", str(data_dir))
    return repo_root, data_dir


def write_excel(path: Path, rows: list[list[object]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


@pytest.mark.asyncio
async def test_generate_verify_builds_prompt_from_artifact_file(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)
    material_dir = tmp_path / "verify-material"
    material_dir.mkdir()
    artifact_path = tmp_path / "营业证照--要素提示词.json"
    artifact_path.write_text(
        """
        {
          "version": "1",
          "carriername": "营业证照",
          "template": {"prompt_template": "HEAD\\n$(factors)\\nTAIL"},
          "factors": [
            {
              "index": 1,
              "factorname": "统一社会信用代码",
              "factortype": "1",
              "factoruse": "企业识别",
              "factor_prompt": "识别18位字母数字组合",
              "source": "ai_generated"
            }
          ]
        }
        """.strip(),
        encoding="utf-8",
    )

    captured = {}

    def fake_run_verify_extraction(material_dir_arg, prompt_text, model, api_key, base_url="", extra_params=None, llm_logs=None):
        captured["material_dir"] = material_dir_arg
        captured["prompt_text"] = prompt_text
        captured["model"] = model
        return {
            "success": True,
            "image_file": "sample.jpg",
            "extraction_output": '{"msginfo":[]}',
            "error": "",
        }

    monkeypatch.setattr(main, "run_verify_extraction", fake_run_verify_extraction)

    expected_prompt = build_preview_prompt(
        "HEAD\n$(factors)\nTAIL",
        [
            {
                "factorname": "统一社会信用代码",
                "factor_prompt": "识别18位字母数字组合",
            }
        ],
    )

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)
            response = await client.post(
                "/api/generate/verify",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "materialDir": str(material_dir),
                    "promptText": "should be ignored",
                    "artifactFile": str(artifact_path),
                },
            )

    assert response.status_code == 200
    assert response.json()["data"]["success"] is True
    assert captured["material_dir"] == material_dir
    assert captured["prompt_text"] == expected_prompt
    assert captured["prompt_text"] != "should be ignored"


@pytest.mark.asyncio
async def test_generate_verify_builds_prompt_from_inline_artifact(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)
    material_dir = tmp_path / "verify-inline"
    material_dir.mkdir()

    captured = {}

    def fake_run_verify_extraction(material_dir_arg, prompt_text, model, api_key, base_url="", extra_params=None, llm_logs=None):
        captured["material_dir"] = material_dir_arg
        captured["prompt_text"] = prompt_text
        return {
            "success": True,
            "image_file": "sample.jpg",
            "extraction_output": '{"msginfo":[]}',
            "error": "",
        }

    monkeypatch.setattr(main, "run_verify_extraction", fake_run_verify_extraction)

    artifact = {
        "version": "1",
        "carriername": "营业证照",
        "template": {"prompt_template": "HEAD\n$(factors)\nTAIL"},
        "factors": [
            {
                "index": 1,
                "factorname": "企业名称",
                "factortype": "1",
                "factoruse": "企业全称",
                "factor_prompt": "识别完整企业名称",
                "source": "manual_edit",
            }
        ],
    }

    expected_prompt = build_preview_prompt(
        artifact["template"]["prompt_template"],
        artifact["factors"],
    )

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)
            response = await client.post(
                "/api/generate/verify",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "materialDir": str(material_dir),
                    "promptText": "fallback text",
                    "artifact": artifact,
                },
            )

    assert response.status_code == 200
    assert response.json()["data"]["success"] is True
    assert captured["material_dir"] == material_dir
    assert captured["prompt_text"] == expected_prompt
    assert captured["prompt_text"] != "fallback text"


@pytest.mark.asyncio
async def test_save_artifact_persists_normalized_json_and_preview(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)
    artifact_path = tmp_path / "营业证照--要素提示词.json"
    preview_path = tmp_path / "营业证照--要素提取完整提示词.txt"

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)
            response = await client.post(
                "/api/generate/save-artifact",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "filePath": str(artifact_path),
                    "previewFilePath": str(preview_path),
                    "artifact": {
                        "carriername": "营业证照",
                        "template": {"prompt_template": "HEAD\n$(factors)\nTAIL"},
                        "factors": [
                            {
                                "name": "统一社会信用代码",
                                "factoruse": "企业识别",
                                "rule": "识别18位字母数字组合",
                                "format": "保持原样",
                            }
                        ],
                    },
                },
            )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    saved_artifact = payload["data"]["artifact"]
    assert saved_artifact["factors"][0]["factorname"] == "统一社会信用代码"
    assert saved_artifact["factors"][0]["factor_prompt"] == "识别18位字母数字组合，保持原样"
    assert saved_artifact["factors"][0]["source"] == "manual"

    saved_json = artifact_path.read_text(encoding="utf-8")
    preview_text = preview_path.read_text(encoding="utf-8")
    assert "统一社会信用代码" in saved_json
    assert "识别18位字母数字组合，保持原样" in saved_json
    assert "HEAD" in preview_text
    assert "## 1.统一社会信用代码" in preview_text
    assert "TAIL" in preview_text


@pytest.mark.asyncio
async def test_generate_download_result_bundles_generated_artifacts(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "generate-download"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])
            material_dir = work_dir / "营业证照"
            material_dir.mkdir(parents=True, exist_ok=True)
            write_excel(
                work_dir / "factors.xlsx",
                [
                    ["材料名称", "要素字段名称", "要素提取说明", "审查要点名称", "审查要点规则说明"],
                    ["营业证照", "统一社会信用代码", "企业唯一识别码", "统一社会信用代码校验", "#营业证照-统一社会信用代码#不能为空"],
                ],
            )
            (material_dir / "营业证照--要素提取完整提示词.txt").write_text("prompt", encoding="utf-8")
            (material_dir / "营业证照--要素提示词.json").write_text('{"ok":true}', encoding="utf-8")
            (material_dir / "营业证照--要素信息录入.json").write_text('{"factors":[]}', encoding="utf-8")

            response = await client.get(
                "/api/generate/download-result",
                headers={"Authorization": f"Bearer {token}"},
                params={"workDir": str(work_dir)},
            )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/zip")

    with ZipFile(io.BytesIO(response.content)) as zf:
        names = set(zf.namelist())

    assert "营业证照/营业证照--要素提取完整提示词.txt" in names
    assert "营业证照/营业证照--要素提示词.json" in names
    assert "营业证照/营业证照--要素信息录入.json" in names
    assert "factors.xlsx" in names


@pytest.mark.asyncio
async def test_factors_workbook_routes_support_online_repair_and_revalidation(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-repair"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])
            material_dir = work_dir / "营业证照"
            material_dir.mkdir(parents=True, exist_ok=True)
            (material_dir / "sample.jpg").write_bytes(b"img")
            write_excel(
                work_dir / "factors.xlsx",
                [
                    ["材料名称", "要素字段名称", "要素提取说明"],
                    ["营业证照", "统一社会信用代码", "企业唯一识别码"],
                ],
            )

            validation_response = await client.post(
                "/api/generate/validate-factors",
                headers={"Authorization": f"Bearer {token}"},
                json={"workDir": str(work_dir), "materials": ["营业证照"]},
            )
            validation_payload = validation_response.json()["data"]
            assert validation_payload["ok"] is False
            assert any(item.get("column") == "审查要点名称" for item in validation_payload.get("diagnostics", []))
            assert any(item.get("column") == "审查要点规则说明" for item in validation_payload.get("diagnostics", []))

            workbook_response = await client.get(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                params={"workDir": str(work_dir)},
            )
            workbook_payload = workbook_response.json()["data"]
            assert workbook_payload["headers"] == ["材料名称", "要素字段名称", "要素提取说明"]
            assert workbook_payload["rows"][0]["values"] == ["营业证照", "统一社会信用代码", "企业唯一识别码"]

            save_response = await client.put(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "workDir": str(work_dir),
                    "headers": [
                        "材料名称",
                        "要素字段名称",
                        "要素提取说明",
                        "审查要点名称",
                        "审查要点规则说明",
                    ],
                    "rows": [
                        {
                            "values": [
                                "营业证照",
                                "统一社会信用代码",
                                "企业唯一识别码",
                                "统一社会信用代码校验",
                                "#营业证照-统一社会信用代码#不能为空",
                            ]
                        }
                    ],
                },
            )

            assert save_response.status_code == 200
            save_payload = save_response.json()["data"]
            assert save_payload["validation"]["ok"] is True
            assert save_payload["validation"]["errors"] == []
            assert save_payload["workbook"]["headers"] == [
                "材料名称",
                "要素字段名称",
                "要素提取说明",
                "审查要点名称",
                "审查要点规则说明",
            ]
            assert save_payload["workbook"]["rows"][0]["values"][3:] == [
                "统一社会信用代码校验",
                "#营业证照-统一社会信用代码#不能为空",
            ]


@pytest.mark.asyncio
async def test_factors_workbook_save_handles_merged_cells(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-merged-cells"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])

            workbook_path = work_dir / "factors.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["事项名称", "材料名称", "要素字段名称", "审查要点规则说明"])
            ws.append(["道路运输证申领", "道路运输证申领登记表", "业户名称", "需与#营业证照-统一社会信用代码#一致"])
            ws.append(["道路运输证申领", None, "车辆号牌", "需与#机动车行驶证-号牌号码#一致"])
            ws.merge_cells("A2:A3")
            ws.merge_cells("B2:B3")
            wb.save(workbook_path)
            wb.close()

            load_response = await client.get(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                params={"workDir": str(work_dir)},
            )

            assert load_response.status_code == 200
            load_payload = load_response.json()["data"]
            assert load_payload["headers"] == ["事项名称", "材料名称", "要素字段名称", "审查要点规则说明"]
            assert [row["rowNumber"] for row in load_payload["rows"]] == [2, 3]
            assert load_payload["rows"][0]["values"] == [
                "道路运输证申领",
                "道路运输证申领登记表",
                "业户名称",
                "需与#营业证照-统一社会信用代码#一致",
            ]
            assert load_payload["rows"][1]["values"] == [
                "道路运输证申领",
                "道路运输证申领登记表",
                "车辆号牌",
                "需与#机动车行驶证-号牌号码#一致",
            ]
            assert load_payload["mergedRanges"] == [
                {"startRow": 2, "endRow": 3, "startColumn": 1, "endColumn": 1},
                {"startRow": 2, "endRow": 3, "startColumn": 2, "endColumn": 2},
            ]

            save_response = await client.put(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "workDir": str(work_dir),
                    "headers": ["事项名称", "材料名称", "要素字段名称", "审查要点规则说明"],
                    "mergedRanges": load_payload["mergedRanges"],
                    "rows": [
                        {
                            "rowNumber": 2,
                            "values": [
                                "道路运输证申领",
                                "道路运输证申领登记表",
                                "业户名称",
                                "需与#营业证照-统一社会信用代码#一致",
                            ]
                        },
                        {
                            "rowNumber": 3,
                            "values": [
                                "道路运输证申领",
                                "道路运输证申领登记表",
                                "车辆号牌",
                                "需与#机动车行驶证-号牌号码#一致",
                            ]
                        },
                    ],
                },
            )

    assert save_response.status_code == 200
    payload = save_response.json()["data"]
    assert payload["workbook"]["rows"][0]["values"] == [
        "道路运输证申领",
        "道路运输证申领登记表",
        "业户名称",
        "需与#营业证照-统一社会信用代码#一致",
    ]
    assert payload["workbook"]["rows"][1]["values"] == [
        "道路运输证申领",
        "道路运输证申领登记表",
        "车辆号牌",
        "需与#机动车行驶证-号牌号码#一致",
    ]
    assert payload["workbook"]["mergedRanges"] == [
        {"startRow": 2, "endRow": 3, "startColumn": 1, "endColumn": 1},
        {"startRow": 2, "endRow": 3, "startColumn": 2, "endColumn": 2},
    ]

    persisted = openpyxl.load_workbook(workbook_path)
    try:
        worksheet = persisted.active
        assert sorted(str(item) for item in worksheet.merged_cells.ranges) == ["A2:A3", "B2:B3"]
        assert worksheet["A2"].value == "道路运输证申领"
        assert worksheet["B2"].value == "道路运输证申领登记表"
        assert worksheet["C3"].value == "车辆号牌"
    finally:
        persisted.close()


@pytest.mark.asyncio
async def test_factors_workbook_load_keeps_blank_rows_inside_merged_regions(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-merged-blank-rows"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])

            workbook_path = work_dir / "factors.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["事项名称", "材料名称", "要素字段名称", "审查要点规则说明"])
            ws.append(["道路运输证申领", "道路运输证申领登记表", "业户名称", "需与#营业证照-统一社会信用代码#一致"])
            ws.append([None, None, None, None])
            ws.append([None, None, "车辆号牌", "需与#机动车行驶证-号牌号码#一致"])
            ws.merge_cells("A2:A4")
            ws.merge_cells("B2:B4")
            wb.save(workbook_path)
            wb.close()

            response = await client.get(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                params={"workDir": str(work_dir)},
            )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert [row["rowNumber"] for row in payload["rows"]] == [2, 3, 4]
    assert payload["rows"][1]["values"] == [
        "道路运输证申领",
        "道路运输证申领登记表",
        "",
        "",
    ]
    assert payload["rows"][2]["values"] == [
        "道路运输证申领",
        "道路运输证申领登记表",
        "车辆号牌",
        "需与#机动车行驶证-号牌号码#一致",
    ]


@pytest.mark.asyncio
async def test_factors_workbook_repair_suggestions_route_returns_confirmable_suggestions(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-repair-suggestions"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])

            response = await client.post(
                "/api/workspaces/factors-workbook/repair-suggestions",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "workDir": str(work_dir),
                    "useLlm": False,
                    "headers": ["材料名称", "要素字段名称", "审查要点名称", "审查要点规则说明"],
                    "rows": [
                        {
                            "rowNumber": 2,
                            "values": ["营业证照", "统一社会信用代码", "", ""],
                        },
                        {
                            "rowNumber": 3,
                            "values": ["营业证照", "", "统一社会信用代码校验", "#营业证照-统一社会信用代吗#不能为空"],
                        },
                    ],
                    "diagnostics": [
                        {
                            "id": "diagnostic-1",
                            "code": "missing_referenced_factor",
                            "row": 3,
                            "column": "审查要点规则说明",
                            "materialName": "营业证照",
                            "factorName": "统一社会信用代吗",
                            "token": "营业证照-统一社会信用代吗",
                            "message": "引用的要素不存在",
                        }
                    ],
                },
            )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["stats"]["total"] == 1
    assert payload["stats"]["suggested"] == 1
    assert payload["items"][0]["diagnosticId"] == "diagnostic-1"
    assert payload["items"][0]["requiresConfirmation"] is True
    assert payload["items"][0]["patches"] == [
        {
            "type": "cell_update",
            "rowNumber": 3,
            "columnIndex": 3,
            "before": "#营业证照-统一社会信用代吗#不能为空",
            "after": "#营业证照-统一社会信用代码#不能为空",
        }
    ]


@pytest.mark.asyncio
async def test_factors_workbook_apply_repair_suggestion_route_can_clone_workspace_material_files(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-apply-repair-suggestion"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])

            source_dir = work_dir / "营业执照"
            source_dir.mkdir(parents=True, exist_ok=True)
            (source_dir / "sample-1.jpg").write_bytes(b"fake-image-data")

            workbook_path = work_dir / "factors.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["事项名称", "材料名称", "要素字段名称", "审查要点名称", "审查要点规则说明"])
            ws.append(["道路运输证申领", "营业执照", "统一社会信用代码", "", ""])
            ws.append(["道路运输证申领", "营业证照", "统一社会信用代码", "", ""])
            wb.save(workbook_path)
            wb.close()

            response = await client.post(
                "/api/workspaces/factors-workbook/apply-repair-suggestion",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "workDir": str(work_dir),
                    "patches": [
                        {
                            "type": "workspace_material_clone",
                            "sourceMaterialName": "营业执照",
                            "targetMaterialName": "营业证照",
                            "sourceFileCount": 1,
                        }
                    ],
                },
            )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["applied"]["count"] == 1
    assert payload["validation"]["ok"] is False
    assert all(item.get("code") != "missing_material_directory" for item in payload["validation"]["diagnostics"])
    assert (work_dir / "营业证照" / "sample-1.jpg").exists()


@pytest.mark.asyncio
async def test_factors_workbook_load_carries_forward_group_headers_without_merge(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-carry-forward-group-columns"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])

            workbook_path = work_dir / "factors.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["事项名称", "材料名称", "要素字段名称", "审查要点规则说明"])
            ws.append(["道路运输证申领", "道路运输证申领登记表", "业户名称", "需与#营业证照-统一社会信用代码#一致"])
            ws.append([None, None, "车辆号牌", "需与#机动车行驶证-号牌号码#一致"])
            ws.append([None, None, "经营许可证号", ""])
            wb.save(workbook_path)
            wb.close()

            response = await client.get(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                params={"workDir": str(work_dir)},
            )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert [row["rowNumber"] for row in payload["rows"]] == [2, 3, 4]
    assert payload["rows"][1]["values"] == [
        "道路运输证申领",
        "道路运输证申领登记表",
        "车辆号牌",
        "需与#机动车行驶证-号牌号码#一致",
    ]
    assert payload["rows"][2]["values"] == [
        "道路运输证申领",
        "道路运输证申领登记表",
        "经营许可证号",
        "",
    ]
    assert payload["mergedRanges"] == []


@pytest.mark.asyncio
async def test_factors_workbook_download_returns_latest_saved_file_with_no_cache_headers(tmp_path, monkeypatch):
    _bootstrap_env(tmp_path, monkeypatch)

    transport = httpx.ASGITransport(app=app)
    async with app.router.lifespan_context(app):
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            token = await login(client)

            workspace_resp = await client.post(
                "/api/workspaces",
                headers={"Authorization": f"Bearer {token}"},
                data={"name": "factors-download-latest"},
            )
            workspace = workspace_resp.json()
            work_dir = Path(workspace["rootPath"])

            workbook_path = work_dir / "factors.xlsx"
            write_excel(
                workbook_path,
                [
                    ["材料名称", "要素字段名称", "要素提取说明"],
                    ["营业证照", "统一社会信用代码", "旧说明"],
                ],
            )

            save_response = await client.put(
                "/api/workspaces/factors-workbook",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "workDir": str(work_dir),
                    "headers": ["材料名称", "要素字段名称", "要素提取说明"],
                    "rows": [
                        {
                            "rowNumber": 2,
                            "values": ["营业证照", "统一社会信用代码", "新说明"],
                        }
                    ],
                },
            )

            assert save_response.status_code == 200

            download_response = await client.get(
                "/api/files/download",
                headers={"Authorization": f"Bearer {token}"},
                params={"path": str(workbook_path)},
            )

    assert download_response.status_code == 200
    assert download_response.headers["cache-control"] == "no-store, no-cache, must-revalidate, max-age=0"
    assert download_response.headers["pragma"] == "no-cache"
    assert download_response.headers["expires"] == "0"

    downloaded = openpyxl.load_workbook(io.BytesIO(download_response.content))
    try:
        worksheet = downloaded.active
        assert worksheet["C2"].value == "新说明"
    finally:
        downloaded.close()
